"""
NUOVO 2026-08-02 — task #172: il foglio di calcolo della valigia.

Richiesta di Lorenzo, alla lettera:

  "per quanto riguarda la parte della valigia mi piace molto ma dopo l'elenco
   vorrei che creassi un collegamento per un foglio di calcolo google come
   quello che ti ho allegato ovviamente costruito in base a ciò che richiede
   la valigia, ma strutturato in maniera simile a quello allegato"

COSA STIAMO RISOLVENDO DAVVERO
------------------------------
Il PDF è un documento che si LEGGE. Una lista della valigia non si legge: si
SPUNTA, in tre momenti diversi, spesso da due persone che preparano la stessa
valigia in due stanze diverse. Su carta la si spunta una volta sola e con una
penna; su un foglio condiviso la si spunta in due, dal telefono, e chi arriva
secondo vede cosa manca. È la differenza fra un contenuto e uno strumento, e
Lorenzo l'ha vista guardando il foglio che usa lui per i suoi viaggi.

PERCHÉ UN FILE `.xlsx` E NON UN FOGLIO GOOGLE CREATO DA NOI (per ora)
----------------------------------------------------------------------
Creare un foglio DENTRO il Drive di qualcuno richiede un account di servizio
Google, credenziali nuove da custodire, e una cartella che continua a
riempirsi di fogli di clienti — cioè dati personali che oggi non conserviamo
e che, non conservandoli, non possiamo perdere.

Un `.xlsx` allegato alla stessa mail del PDF ottiene la stessa cosa senza
niente di tutto ciò: il cliente lo apre e, se vuole il foglio condiviso, in
Google Drive fa "Apri con Fogli Google" e da lì lo condivide con chi vuole.
Funziona anche per chi Google non lo usa (Excel, Numbers, LibreOffice) — e i
clienti che non usano Google non sono pochi.

La strada verso il foglio creato da noi resta APERTA e già cablata: se un
domani `CHECKLIST_SHEET_TEMPLATE_URL` viene configurato, il riquadro nel PDF
punta lì invece che al file allegato, senza toccare una riga di codice. È la
scelta "doppio binario" presa con Lorenzo il 2026-08-02.

DA DOVE VENGONO LE RIGHE
------------------------
Da NESSUNA fonte nuova. Il foglio è una seconda vista degli stessi dati che
generano il capitolo della valigia e la lista della sera prima:

    vademecum["packing"]   → le voci per gruppo
    vademecum["baggage"]   → il tipo di bagaglio e il suo costo
    vademecum["climate"]   → il link alla previsione vera, e QUANDO aprirlo
    predeparture["checklist"] → le cose che se mancano ti bloccano

Questo è deliberato e non è pigrizia: se il foglio avesse una sorgente
propria, il giorno in cui il capitolo cambia il foglio direbbe un'altra cosa,
e il cliente si troverebbe due liste in disaccordo dentro la stessa mail.
Nessuna chiamata di rete, nessun token, nessuna latenza aggiunta.

LA STRUTTURA, PRESA DAL FOGLIO DI LORENZO
------------------------------------------
Stesse colonne e stessa idea di fasce colorate del suo `Checklist_Viaggio`:
priorità in bande, "quando", categoria, cosa fare, le caselle da spuntare
(una per viaggiatore), il collegamento e le note. Due differenze, entrambe
volute:

  1. le fasce non dicono "entro 2 giorni" ma la DATA VERA, calcolata dalla
     partenza di questo viaggio. "Entro il 9 settembre" non si interpreta;
     "entro 2 giorni" da quando?
  2. le caselle sono celle booleane vere (non la scritta "FALSE"): è ciò che
     fa comparire la spunta cliccabile quando il file viene aperto in Fogli
     Google, ed è esattamente come è fatto il foglio allegato da Lorenzo.

[AGGIUNTO 2026-08-03] UNA CASELLA A TESTA, E LA STRADA DI RITORNO
------------------------------------------------------------------
Due richieste di Lorenzo, alla lettera: "ricordati di aggiungere poi le spunte
per i viaggiatori (se sono tre, 3 caselle di checklist, se sono 4 ne metti 4 e
così via) e ovviamente un pulsante sul foglio di calcolo che ti fa ritornare al
pdf originario".

  * UNA COLONNA PER VIAGGIATORE VERO. Prima erano al massimo quattro: il quinto
    di un gruppo di cinque non aveva dove spuntare, e un foglio dove uno dei
    cinque non può spuntare torna a essere una lista da leggere. Il tetto resta
    ma sale (`MAX_COLONNE_SPUNTA`), e quando taglia LO DICE nel foglio: il
    perché del numero, e i tempi misurati che lo giustificano, stanno accanto
    alla costante.

  * IL PULSANTE DI RITORNO (`itinerary_url`). Facoltativo, e vale solo se è un
    indirizzo `https://` vero: senza indirizzo usabile non compare NESSUN
    pulsante, mai una casella che promette il ritorno all'itinerario e non apre
    niente. È la stessa regola del riquadro nel PDF, che ripiega sul nome
    dell'allegato quando `CHECKLIST_SHEET_TEMPLATE_URL` non è configurata.
    Finché nessun chiamante passa l'indirizzo, il foglio esce identico a prima:
    l'intestazione resta alla riga 1 e non si sposta niente.
"""
from __future__ import annotations

import io
import math
import re
from datetime import date, timedelta

from src import identita

MESI = (
    "", "gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
    "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre",
)

# Le bande, nell'ordine in cui il foglio va letto dall'alto. `giorni_prima` è
# quanti giorni PRIMA della partenza si chiude quella banda; `None` = durante.
# I numeri non sono inventati: 30 giorni è la soglia sotto la quale rinnovare
# un documento d'identità non è più realistico in Italia, 14 è la finestra in
# cui i biglietti a orario dei musei più visitati si esauriscono, 3 è il primo
# giorno in cui una previsione meteo esiste davvero (vedi `vademecum.py`).
BANDE = (
    {"key": "subito", "label": "🔴 SUBITO", "breve": "Subito",
     "giorni_prima": 30, "fill": identita.excel(identita.FASCE["subito"])},
    {"key": "due_settimane", "label": "🟠 DUE SETTIMANE PRIMA",
     "breve": "2 settimane", "giorni_prima": 14,
     "fill": identita.excel(identita.FASCE["due_settimane"])},
    {"key": "settimana", "label": "🟡 UNA SETTIMANA PRIMA",
     "breve": "1 settimana", "giorni_prima": 7,
     "fill": identita.excel(identita.FASCE["settimana"])},
    {"key": "vigilia", "label": "🟢 IL GIORNO PRIMA", "breve": "Vigilia",
     "giorni_prima": 1, "fill": identita.excel(identita.FASCE["vigilia"])},
    {"key": "viaggio", "label": "🔵 DURANTE IL VIAGGIO",
     "breve": "In viaggio", "giorni_prima": None,
     "fill": identita.excel(identita.FASCE["viaggio"])},
)
_BANDA_INDICE = {b["key"]: i for i, b in enumerate(BANDE)}

# [RIFATTO 2026-08-05 — task #193/#194] I colori vengono tutti da
# `src/identita.py`. Prima erano scritti qui a mano, ed erano i verdi e i
# gialli predefiniti di Fogli Google: il foglio sembrava un foglio, non un
# pezzo del prodotto. Adesso la tavolozza è una sola per il PDF, per i
# capitoli staccati e per il foglio — cambiare un colore in un posto lo
# cambia dappertutto, che è tutta la differenza fra avere un'identità e
# avere dei colori.
HEADER_FILL = identita.excel(identita.NOTTE)

# [MISURATO 2026-08-03 — richiesta di Lorenzo: "se sono tre, 3 caselle di
# checklist, se sono 4 ne metti 4 e così via"]
# Una colonna per viaggiatore VERO, senza il vecchio tetto a 4: il quinto di un
# gruppo di cinque non aveva dove spuntare, e una lista dove uno dei cinque non
# può spuntare torna a essere una lista da leggere.
#
# Il tetto però resta, e non per estetica: `travellers` arriva dal modulo di
# richiesta e passa per Make, quindi può essere 10.000 per un errore di
# battitura. Misurato scrivendo lo stesso foglio (32 righe) al variare delle
# colonne, su questa macchina:
#     12 colonne →  0,02 s,  6,8 KB, 372 caratteri di larghezza totale
#     50 colonne →  0,13 s, 10,2 KB
#    200 colonne →  1,41 s
#   1000 colonne → 23,7  s, 102 KB
#  16380 colonne → non finito dopo DIECI MINUTI (prova interrotta)
# Oltre poche centinaia di colonne il problema non è più un foglio brutto: è la
# generazione che sfonda il tetto dei 300 secondi di Make e porta via il PDF a
# tutti, non solo il foglio.
#
# Perché 12 e non 8 o 20: 12 è il gruppo più grande che condivide DAVVERO una
# valigia e un itinerario (due famiglie che partono insieme, o un nove posti più
# una seconda auto); sopra, è un viaggio di gruppo con un organizzatore che ha i
# suoi strumenti. Ed è anche il limite di leggibilità: 12 colonne di spunta sono
# 168 caratteri di larghezza, che con le quattro colonne fisse fanno 372 — circa
# due schermate in orizzontale. A 20 sarebbero 484, cioè tre.
MAX_COLONNE_SPUNTA = 12

# Il pulsante che riporta al PDF (richiesta di Lorenzo: "un pulsante sul foglio
# di calcolo che ti fa ritornare al pdf originario"). Il testo sta qui e non
# dentro la funzione perché è la stessa identica cosa sui due fogli, ed è ciò
# che i test cercano per sapere se il pulsante c'è.
TESTO_BOTTONE_ITINERARIO = "↩ Torna all'itinerario"
BOTTONE_FILL = identita.excel(identita.NOTTE)


# ---------------------------------------------------------------------------
# Lettura difensiva degli ingredienti
# ---------------------------------------------------------------------------
def _get(obj, name, default=None):
    if obj is None:
        return default
    if isinstance(obj, dict):
        return obj.get(name, default)
    return getattr(obj, name, default)


def _parse_date(value) -> date | None:
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        return None
    try:
        return date.fromisoformat(value.strip()[:10])
    except ValueError:
        return None


def _data_leggibile(giorno: date) -> str:
    return f"{giorno.day} {MESI[giorno.month]}"


def _quando(partenza: date | None, giorni_prima: int | None) -> str:
    """La colonna "Quando", con la data vera quando la conosciamo.

    Senza data di partenza si scrive comunque qualcosa di utile ("30 giorni
    prima di partire"): meglio una regola generica che una cella vuota, e
    molto meglio di una data inventata.
    """
    if giorni_prima is None:
        return "Sul posto"
    if partenza is None:
        return f"{giorni_prima} giorni prima di partire" if giorni_prima > 1 else "Il giorno prima"
    scadenza = partenza - timedelta(days=giorni_prima)
    if scadenza <= date.today():
        return f"Subito (la partenza è il {_data_leggibile(partenza)})"
    return f"Entro il {_data_leggibile(scadenza)}"


# ---------------------------------------------------------------------------
# Da quale gruppo della valigia nasce quale banda e quale categoria
# ---------------------------------------------------------------------------
# La classificazione guarda il NOME del gruppo prodotto da
# `vademecum.build_packing()`. Un gruppo che domani cambiasse nome non
# sparirebbe: cadrebbe nel ramo predefinito ("Valigia", banda "una settimana
# prima"), che è la collocazione giusta per una voce di valigia qualsiasi.
def _classifica_gruppo(nome_gruppo: str) -> tuple[str, str]:
    testo = (nome_gruppo or "").lower()
    if "documenti" in testo or "salute" in testo:
        # I documenti stanno in cima per un motivo pratico: sono le uniche
        # voci che, se mancano, non si rimediano il giorno prima.
        return "subito", "Documenti"
    if "elettronica" in testo:
        return "settimana", "Elettronica"
    if "clima" in testo:
        return "settimana", "Abbigliamento"
    if "programma" in testo:
        return "due_settimane", "Programma"
    if "quanto portare" in testo:
        return "vigilia", "Valigia"
    return "settimana", "Valigia"


_PRIMA_FRASE = re.compile(r"^(.{8,90}?)(?:[:.]\s|\s—\s|$)")


def _titolo_e_nota(voce: str) -> tuple[str, str]:
    """Spezza una voce lunga in "cosa fare" + "perché".

    Le voci della valigia sono scritte per essere LETTE in un paragrafo, e
    sono lunghe apposta: la ragione accanto alla cosa è ciò che le rende
    convincenti. In una tabella da spuntare, però, una cella di quaranta
    parole rende la riga illeggibile e la colonna della spunta lontanissima.
    Quindi la ragione non si butta: si sposta in "Note", dove resta a
    disposizione di chi la vuole senza intasare la lista.
    """
    testo = (voce or "").strip()
    if not testo:
        return "", ""
    match = _PRIMA_FRASE.match(testo)
    if not match:
        return testo, ""
    titolo = match.group(1).strip(" :.—-")
    resto = testo[match.end(1):].lstrip(" :.—-").strip()
    if not titolo:
        return testo, ""
    return titolo, resto


# ---------------------------------------------------------------------------
# Le righe
# ---------------------------------------------------------------------------
def build_checklist_rows(trip=None, vademecum: dict | None = None,
                         predeparture: dict | None = None) -> list[dict]:
    """Le righe del foglio: `[{"banda", "quando", "categoria", "attivita",
    "link", "note"}]`, già nell'ordine di lettura.

    Non solleva mai: un ingrediente mancante toglie le sue righe, non il
    foglio. Un foglio con dieci righe è utile; un allegato che non si apre è
    un difetto che il cliente vede prima di ogni altra cosa.
    """
    partenza = _parse_date(_get(trip, "date_start"))
    righe: list[dict] = []

    # [MISURATO, non previsto] Alla prima prova il foglio usciva con
    # "Documento d'identita' valido..." DUE volte: una dalla lista della sera
    # prima e una dal gruppo "Documenti e salute" della valigia. Nel PDF non
    # si nota, perche' sono due capitoli lontani con due scopi diversi; in una
    # tabella da spuntare due righe quasi uguali a otto di distanza fanno
    # perdere fiducia in tutte le altre. Il confronto e' sulle prime parole
    # normalizzate, non sulla frase intera: le due versioni non sono mai
    # identiche carattere per carattere.
    visti: set[str] = set()

    def _impronta(testo: str) -> str:
        parole = re.findall(r"\w+", (testo or "").lower(), flags=re.UNICODE)
        # Quattro parole, non cinque: "Documento d'identita' VALIDO PER tutta
        # la durata" e "Documento d'identita' VALIDO, PIU' una foto" divergono
        # esattamente alla quinta, ed erano proprio il doppione da togliere.
        return " ".join(parole[:4])

    def aggiungi(banda, categoria, attivita, note="", link=""):
        attivita = (attivita or "").strip()
        if not attivita:
            return
        impronta = _impronta(attivita)
        if impronta and impronta in visti:
            return
        visti.add(impronta)
        righe.append({
            "banda": banda,
            "quando": _quando(partenza, BANDE[_BANDA_INDICE[banda]]["giorni_prima"]),
            "categoria": categoria,
            "attivita": attivita,
            "note": (note or "").strip(),
            "link": (link or "").strip(),
        })

    vademecum = vademecum if isinstance(vademecum, dict) else {}
    predeparture = predeparture if isinstance(predeparture, dict) else {}

    # --- Le cose che se mancano ti bloccano: vengono per prime -------------
    for voce in predeparture.get("checklist") or []:
        if not isinstance(voce, dict):
            continue
        aggiungi("subito", "Prima di tutto", voce.get("title"), voce.get("detail"))

    # --- La valigia, gruppo per gruppo ------------------------------------
    for gruppo in vademecum.get("packing") or []:
        if not isinstance(gruppo, dict):
            continue
        banda, categoria = _classifica_gruppo(gruppo.get("group") or "")
        for voce in gruppo.get("items") or []:
            titolo, nota = _titolo_e_nota(voce if isinstance(voce, str) else "")
            aggiungi(banda, categoria, titolo, nota)

    # --- Il bagaglio: una riga sola, ma è quella che costa soldi -----------
    baggage = vademecum.get("baggage")
    if isinstance(baggage, dict):
        scelta = baggage.get("choice")
        if scelta:
            # `reason` dice PERCHE' quella scelta, `total` quanto costa in
            # tutto: sono le due sole cose che servono per decidere, e stanno
            # in una nota sola.
            nota = " ".join(
                x for x in (baggage.get("reason"), baggage.get("total")) if x
            )
            aggiungi(
                "due_settimane", "Bagaglio",
                f"Bagaglio consigliato per questo viaggio: {scelta}",
                nota,
            )
        aggiungi(
            "due_settimane", "Bagaglio",
            "Misura la valigia CON le ruote e il manico, e pesala da piena",
            "I limiti della compagnia si riferiscono all'ingombro totale. È il "
            "controllo che al gate costa più caro a chi non l'ha fatto a casa.",
        )

    # --- Il meteo vero, nel solo momento in cui esiste ---------------------
    climate = vademecum.get("climate")
    if isinstance(climate, dict):
        # `forecast_link` e' un dizionario `{"url", "label"}` (vedi
        # `vademecum.forecast_link`): e' una RICERCA dichiarata come tale, mai
        # un indirizzo di previsione indovinato.
        forecast = climate.get("forecast_link")
        link = forecast.get("url") if isinstance(forecast, dict) else ""

        aggiungi(
            "vigilia", "Meteo",
            "Guarda la previsione vera e correggi la valigia",
            "Prima di tre giorni dalla partenza una previsione meteo non "
            "esiste: quello che si trova è la media del mese travestita da "
            "previsione. Questa riga è il momento giusto per guardarla.",
            link,
        )

    # --- Sul posto ---------------------------------------------------------
    # [MISURATO 2026-08-02] Queste due righe sono vere per OGNI viaggio, e
    # proprio per questo non bastano da sole: senza vademecum e senza lista
    # della sera prima, il cliente si ritroverebbe allegato un foglio di due
    # righe generiche che non parlano del suo viaggio. Meglio nessun allegato
    # che un allegato vuoto: si esce di qui a mani vuote e il riquadro nel PDF
    # non compare (`build_pdf_sections` non riceve righe).
    if not righe:
        return []

    aggiungi("viaggio", "Sul posto",
             "Tieni il PDF dell'itinerario aperto anche senza rete",
             "Salvalo nei file del telefono, non solo nella mail: la mail ha "
             "bisogno di rete, il file no.")
    aggiungi("viaggio", "Sul posto",
             "Spunta qui quello che NON hai usato",
             "È l'unica riga di questo foglio che serve al viaggio dopo: la "
             "valigia si alleggerisce solo con le prove, non con i propositi.")

    righe.sort(key=lambda r: _BANDA_INDICE[r["banda"]])
    return righe


def build_itinerary_rows(itinerary: dict | None = None, trip=None) -> list[dict]:
    """Il secondo foglio: `[{"data", "luogo", "programma"}]`.

    Serve a chi condivide il foglio con chi parte insieme a lui e non ha il
    PDF sotto mano: una riga per giornata, con la data vera e i luoghi in
    ordine. Non sostituisce il programma — non ne avrebbe la ricchezza — ma
    risponde alla sola domanda che si fa dentro un foglio condiviso: "il 12
    dove siamo e cosa facciamo?".
    """
    if not isinstance(itinerary, dict):
        return []
    partenza = _parse_date(_get(trip, "date_start"))
    destinazione = (_get(trip, "destination", "") or "").strip()
    righe: list[dict] = []
    for indice, giornata in enumerate(itinerary.get("days") or []):
        if not isinstance(giornata, dict):
            continue
        giorno = giornata.get("day") or (indice + 1)
        data_giornata = giornata.get("date") or ""
        if not data_giornata and partenza is not None:
            try:
                data_giornata = _data_leggibile(partenza + timedelta(days=int(giorno) - 1))
            except (TypeError, ValueError):
                data_giornata = ""
        else:
            parsed = _parse_date(data_giornata)
            if parsed is not None:
                data_giornata = _data_leggibile(parsed)
        tappe = []
        for blocco in giornata.get("blocks") or giornata.get("stops") or []:
            if not isinstance(blocco, dict):
                continue
            # [MISURATO] Il campo giusto e' `activity` ("Salita alla Torre del
            # Mangia"), non `location`: quest'ultimo e' l'indirizzo ("Via
            # Giovanni Dupre' 132"), e una riga fatta di indirizzi non si
            # riconosce nemmeno da chi c'e' stato.
            nome = (
                blocco.get("activity") or blocco.get("poi_name")
                or blocco.get("name") or blocco.get("location")
            )
            if not nome:
                continue
            ora = str(blocco.get("time") or "").strip()
            tappe.append(f"{ora} {nome}".strip())
        righe.append({
            "data": f"Giorno {giorno}" + (f" · {data_giornata}" if data_giornata else ""),
            "luogo": giornata.get("title") or destinazione,
            "programma": " → ".join(tappe),
        })
    return righe


# ---------------------------------------------------------------------------
# Il file
# ---------------------------------------------------------------------------
def _quante_spunte(travellers) -> tuple[int, int | None]:
    """`(colonne da fare, numero chiesto se è stato tagliato)`.

    Il secondo valore non serve al foglio: serve a POTERLO DIRE. Tagliare in
    silenzio da 20 a 12 lascia chi ha scritto "siamo in 20" a contare colonne
    e a chiedersi se il suo conto è arrivato o se il file è rotto.

    Non solleva per nessun ingresso. `travellers` attraversa il modulo di
    richiesta e Make: arriva anche come `"due"`, `None`, `2.5` o vuoto.
    """
    try:
        # `float` prima di `ceil` perché la stringa "3" deve valere 3 come il
        # numero 3 — dal modulo arriva nei due modi a seconda di come Make
        # serializza il campo. `nan` e `inf` cadono qui sotto: `ceil` solleva
        # ValueError sul primo e OverflowError sul secondo, e un numero di
        # persone che non è un numero vale 1, non zero colonne.
        chiesti = math.ceil(float(travellers))
    except (TypeError, ValueError, OverflowError):
        return 1, None
    if chiesti > MAX_COLONNE_SPUNTA:
        return MAX_COLONNE_SPUNTA, chiesti
    # Sotto l'1 non si scende: un foglio senza nessuna colonna da spuntare non
    # è un foglio da spuntare.
    return max(1, chiesti), None


def _colonne_spunta(travellers) -> list[str]:
    """Le intestazioni delle colonne da spuntare, una per viaggiatore.

    Nel foglio di Lorenzo sono i nomi delle persone ("check Luca", "check
    Debby"). Noi i nomi non li abbiamo e non li chiediamo: chiederli
    significherebbe raccogliere dati di persone che non sono nostri clienti.
    Quindi una colonna per viaggiatore, numerata, che chiunque rinomina in due
    secondi con il nome vero appena apre il foglio.
    """
    quanti, _ = _quante_spunte(travellers)
    if quanti == 1:
        return ["Fatto"]
    return [f"Fatto · viaggiatore {i}" for i in range(1, quanti + 1)]


# Un indirizzo entra nel foglio solo se è assoluto, cifrato e con un host vero.
# `https://` da solo, `//esempio/x` e `non-un-url` sembrano indirizzi e non
# aprono niente: diventerebbero un pulsante morto, che costa più fiducia di
# quanta ne dia il pulsante che non c'è. `http://` è escluso come ovunque nel
# progetto. Gli spazi dentro l'indirizzo lo rendono inutilizzabile in un
# collegamento, quindi lo scartano.
_URL_ITINERARIO = re.compile(r"^https://[^\s/]+\.[^\s/]+(?:/\S*)?$", re.IGNORECASE)


def _url_itinerario(itinerary_url) -> str:
    """L'indirizzo del PDF se è usabile davvero, altrimenti `""` (= niente
    pulsante). Non solleva: un tipo sbagliato è solo un indirizzo assente."""
    if not isinstance(itinerary_url, str):
        return ""
    pulito = itinerary_url.strip()
    return pulito if _URL_ITINERARIO.match(pulito) else ""


def build_checklist_filename(trip=None) -> str:
    """Il nome del file come lo vede il cliente nella mail.

    Un allegato che si chiama `checklist.xlsx` in una casella con altri
    quattro allegati è un file che non si ritrova. Con destinazione e mese
    dentro il nome, si ritrova anche fra un anno.
    """
    destinazione = (_get(trip, "destination", "") or "").strip()
    pezzi = ["Valigia"]
    if destinazione:
        pulita = re.sub(r"[^\w\s-]", "", destinazione, flags=re.UNICODE).strip()
        pulita = re.sub(r"\s+", "-", pulita)
        if pulita:
            pezzi.append(pulita)
    partenza = _parse_date(_get(trip, "date_start"))
    if partenza is not None:
        pezzi.append(partenza.strftime("%Y-%m"))
    return "-".join(pezzi) + ".xlsx"


def build_checklist_xlsx(trip=None, vademecum: dict | None = None,
                         predeparture: dict | None = None,
                         itinerary: dict | None = None,
                         travellers: int = 1,
                         itinerary_url: str | None = None) -> bytes | None:
    """Il foglio di calcolo, in memoria. `None` se non c'è niente da spuntare.

    `itinerary_url` è l'indirizzo del PDF dell'itinerario: se è un `https://`
    vero, i due fogli portano in cima il pulsante che ci riporta. Se manca, o
    non è usabile, non compare nessun pulsante — mai una casella che promette
    un ritorno e non apre niente.

    Non solleva mai. Un errore qui non deve poter togliere il PDF: il foglio
    è un di più, il documento è il prodotto.
    """
    try:
        righe = build_checklist_rows(trip, vademecum, predeparture)
    except Exception as e:  # noqa: BLE001
        print(f"⚠️  Foglio valigia: righe non costruite: {type(e).__name__}: {e}")
        return None
    if not righe:
        return None
    try:
        return _scrivi_xlsx(righe, build_itinerary_rows(itinerary, trip), trip,
                            travellers, itinerary_url)
    except Exception as e:  # noqa: BLE001
        print(f"⚠️  Foglio valigia saltato: {type(e).__name__}: {e}")
        return None


def _scrivi_testata(ws, trip, colonne: int) -> None:
    """Il blocco di testa: marchio, destinazione, date.

    [AGGIUNTO 2026-08-05 — task #193, richiesta di Lorenzo: «con quest'ultimo
    svolgi anche un lavoro di miglioramento generale soprattutto di design»]

    Prima il foglio si apriva sulla riga delle intestazioni. Funzionava, e
    sembrava esattamente quello che era: un foglio di calcolo. Chi lo riceve
    lo apre tre settimane dopo aver comprato l'itinerario, spesso dal
    telefono, e la prima cosa che deve capire è di CHE VIAGGIO parla — non
    quali colonne ha.

    Sopralinea in maiuscoletto, titolo grande, riga di contesto sotto: sono
    le stesse tre righe della copertina del PDF, ed è il motivo per cui i due
    documenti si riconoscono come parenti.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    destinazione = (_get(trip, "destination", "") or "").strip()
    partenza = _parse_date(_get(trip, "date_start"))
    ritorno = _parse_date(_get(trip, "date_end"))

    ws.append([identita.MARCHIO])
    riga_marchio = ws.max_row
    ws.append([f"La valigia per {destinazione}" if destinazione else "La valigia"])
    riga_titolo = ws.max_row
    # L'ANNO ci va, e non e' pignoleria: questo foglio si ritrova nei
    # download un anno dopo, accanto a quello del viaggio prima. «12
    # settembre» da solo non dice quale dei due.
    contesto = []
    if partenza is not None:
        contesto.append(f"Partenza {_data_leggibile(partenza)} {partenza.year}")
    if ritorno is not None:
        contesto.append(f"rientro {_data_leggibile(ritorno)} {ritorno.year}")
    ws.append([" \u00b7 ".join(contesto) or "Spunta man mano: il foglio tiene il conto."])
    riga_contesto = ws.max_row

    # Il fondo scuro va su TUTTE le colonne della testata, una per una: le
    # celle non si fondono mai in questo foglio (`openpyxl` 3.1.5 perde il
    # colore sulle celle fuse — vedi la nota sulle fasce). Il testo trabocca
    # sulle celle vuote accanto, che e' l'effetto voluto.
    for indice in (riga_marchio, riga_titolo, riga_contesto):
        for colonna in range(1, colonne + 1):
            ws.cell(row=indice, column=colonna).fill = PatternFill(
                "solid", fgColor=identita.excel(identita.NOTTE))

    ws.cell(row=riga_marchio, column=1).font = Font(
        bold=True, size=9, color=identita.excel(identita.ORO))
    ws.cell(row=riga_titolo, column=1).font = Font(
        bold=True, size=18, color="FFFFFFFF")
    ws.cell(row=riga_contesto, column=1).font = Font(
        size=10, color="FFD8D2C6")
    for indice in (riga_marchio, riga_titolo, riga_contesto):
        ws.cell(row=indice, column=1).alignment = Alignment(vertical="center")
    ws.row_dimensions[riga_marchio].height = 18
    ws.row_dimensions[riga_titolo].height = 30
    ws.row_dimensions[riga_contesto].height = 20


def _scrivi_contatore(ws, colonne: int) -> int:
    """La riga che dice quanto manca. Torna il numero di riga, da riempire dopo.

    [AGGIUNTO 2026-08-05 — task #193] La formula e' VIVA: e' un `CONTA.SE`
    sull'intero blocco delle caselle, quindi il numero cala da solo mentre si
    spunta. Un numero scritto fisso al momento della generazione direbbe
    sempre la stessa cosa e sarebbe peggio di niente — la prima volta che
    resta indietro, chi legge smette di fidarsi anche del resto del foglio.

    La riga si prenota qui e si riempie in fondo, quando si sa dove finisce
    il corpo: e' l'unico ordine possibile, e vale la pena scriverlo perche'
    sembra un giro strano finche' non lo si nota.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.append([""])
    indice = ws.max_row
    for colonna in range(1, colonne + 1):
        ws.cell(row=indice, column=colonna).fill = PatternFill(
            "solid", fgColor=identita.excel(identita.AVORIO))
    cella = ws.cell(row=indice, column=1)
    cella.font = Font(bold=True, size=11, color=identita.excel(identita.NOTTE))
    cella.alignment = Alignment(vertical="center")
    ws.row_dimensions[indice].height = 24
    return indice


def _scrivi_bottone_itinerario(ws, url: str, larghezza_celle: int = 2) -> None:
    """Il pulsante "↩ Torna all'itinerario" in cima al foglio.

    Chi apre il foglio è in piedi davanti alla valigia, con il telefono in
    mano, e il PDF è in una mail di tre settimane fa sotto altre venti. Senza
    questo pulsante la strada dal foglio al documento passa dalla ricerca in
    posta; con il pulsante è un tocco.

    Perché sembra un pulsante e non una riga: fondo colorato, grassetto e
    bordo su tutti i lati. In un foglio di calcolo si clicca solo quello che
    ha l'aria di essere cliccabile, e una cella azzurra in mezzo ad altre celle
    non ce l'ha.

    [MISURATO 2026-08-02, la stessa trappola delle fasce] Le celle NON si
    fondono nemmeno qui: `openpyxl` 3.1.5 accetta il colore sulle celle a
    destra di una fusione e poi non lo salva. Si colorano quindi due celle
    affiancate (larghezza 11+21 = 32 caratteri, contro i 23 del testo): il
    testo trabocca sulla seconda, che è vuota, e il pulsante si vede intero.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    ws.append([TESTO_BOTTONE_ITINERARIO])
    indice = ws.max_row
    # [RIFATTO 2026-08-05 — task #193] Era azzurro chiaro col testo blu
    # sottolineato: il collegamento predefinito di un foglio di calcolo.
    # Adesso e' il blu dell'identita' col testo bianco — lo stesso pulsante
    # che sta dentro i capitoli del PDF. Chi lo vede, l'ha gia' visto.
    bordo = Side(style="thin", color=identita.excel(identita.NOTTE))
    for colonna in range(1, larghezza_celle + 1):
        cella = ws.cell(row=indice, column=colonna)
        cella.fill = PatternFill("solid", fgColor=BOTTONE_FILL)
        cella.font = Font(bold=True, color="FFFFFFFF")
        cella.alignment = Alignment(horizontal="left", vertical="center")
        cella.border = Border(top=bordo, bottom=bordo, left=bordo, right=bordo)
    # Il collegamento sta sulla prima cella: è quella che contiene il testo, ed
    # è l'unica su cui il dito di chi legge finisce davvero.
    ws.cell(row=indice, column=1).hyperlink = url
    ws.row_dimensions[indice].height = 26


def _scrivi_avviso_colonne(ws, chiesti: int, colonne: int) -> None:
    """La riga che dice che le colonne da spuntare sono state tagliate.

    Sta sopra l'intestazione perché è l'unico posto in cui la legge chi conta
    le colonne e non trova il suo numero. Dice tutti e tre i pezzi: quanti ne
    erano stati chiesti, quanti ce ne sono, e cosa fare per gli altri.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.append([
        f"⚠ Viaggiatori indicati: {chiesti}. Questo foglio ha {colonne} colonne "
        f"da spuntare, il massimo che resta leggibile: per gli altri aggiungete "
        f"una colonna a mano, oppure duplicate il foglio."
    ])
    indice = ws.max_row
    cella = ws.cell(row=indice, column=1)
    cella.font = Font(bold=True)
    cella.fill = PatternFill("solid", fgColor="FFFFF2CC")
    cella.alignment = Alignment(vertical="center")


def _scrivi_xlsx(righe: list[dict], righe_itinerario: list[dict],
                 trip, travellers: int, itinerary_url: str | None = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    quante, chiesti_di_troppo = _quante_spunte(travellers)
    spunte = _colonne_spunta(travellers)
    intestazioni = ["Priorità", "Quando", "Categoria", "Attività"] + spunte + ["Link", "Note"]
    url_itinerario = _url_itinerario(itinerary_url)

    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist"

    grassetto = Font(bold=True)
    testata = PatternFill("solid", fgColor=HEADER_FILL)
    a_capo = Alignment(wrap_text=True, vertical="top")
    centrato = Alignment(horizontal="center", vertical="center")
    bordo_leggero = Border(
        bottom=Side(style="thin", color=identita.excel(identita.FILETTO)))
    zebra = PatternFill("solid", fgColor=identita.excel(identita.AVORIO))

    # Sopra l'intestazione ci va SOLO ciò che c'è davvero: niente pulsante se
    # non c'è un indirizzo, niente avviso se non è stato tagliato niente. È il
    # motivo per cui l'intestazione non è a una riga fissa e nel resto della
    # funzione si usa `riga_intestazione` invece del numero 1: senza aggiunte
    # il foglio è identico a prima, riga per riga.
    _scrivi_testata(ws, trip, len(intestazioni))
    if url_itinerario:
        _scrivi_bottone_itinerario(ws, url_itinerario)
    riga_contatore = _scrivi_contatore(ws, len(intestazioni))
    if chiesti_di_troppo:
        _scrivi_avviso_colonne(ws, chiesti_di_troppo, quante)

    ws.append(intestazioni)
    riga_intestazione = ws.max_row
    for cella in ws[riga_intestazione]:
        cella.font = Font(bold=True, color="FFFFFFFF", size=10)
        cella.fill = testata
        cella.alignment = centrato
        cella.border = bordo_leggero
    ws.row_dimensions[riga_intestazione].height = 24

    banda_corrente = None
    prima_riga_corpo = None
    for riga in righe:
        if riga["banda"] != banda_corrente:
            banda_corrente = riga["banda"]
            banda = BANDE[_BANDA_INDICE[banda_corrente]]
            ws.append([banda["label"]])
            indice = ws.max_row
            # [MISURATO 2026-08-02, dal controllo `test_le_fasce_colorate_ci_
            # sono_e_sono_colorate_fino_in_fondo`] La fascia NON va fusa.
            # Sembrava la scelta ovvia — una banda e' una riga sola — ma
            # `openpyxl` 3.1.5, appena si fondono le celle, sostituisce quelle
            # a destra con oggetti `MergedCell` che ACCETTANO il colore e poi
            # non lo salvano: nessun errore, e il file esce con la fascia
            # colorata solo sulla prima colonna e bianca per le altre sette.
            # Provato in tutti e tre i modi (colore prima della fusione, dopo,
            # e sull'intervallo `ws['A2:H2']`): solo senza fusione il colore
            # arriva in fondo. Senza fusione l'etichetta trabocca comunque
            # sulle celle vuote accanto, che e' esattamente l'effetto voluto.
            cella = ws.cell(row=indice, column=1)
            cella.font = Font(bold=True)
            cella.alignment = Alignment(vertical="center")
            for colonna in range(1, len(intestazioni) + 1):
                ws.cell(row=indice, column=colonna).fill = PatternFill(
                    "solid", fgColor=banda["fill"]
                )

        ws.append([
            banda.get("breve") or "", riga["quando"], riga["categoria"],
            riga["attivita"],
            *([None] * len(spunte)),
            riga["link"], riga["note"],
        ])
        indice = ws.max_row
        if prima_riga_corpo is None:
            prima_riga_corpo = indice
        # Zebratura appena percettibile: l'avorio dell'identita', non il
        # grigio. Serve a non perdere la riga scorrendo dodici colonne di
        # caselle, ed e' l'unica decorazione di tutto il foglio.
        for colonna in range(1, len(intestazioni) + 1):
            cella = ws.cell(row=indice, column=colonna)
            cella.alignment = a_capo
            cella.border = bordo_leggero
            if (indice - riga_intestazione) % 2 == 0:
                cella.fill = zebra
        ws.cell(row=indice, column=1).font = Font(
            size=9, bold=True, color=identita.excel(identita.GRIGIO_TESTO))
        # Le caselle da spuntare: celle BOOLEANE vere, non la parola "FALSE".
        # È questo che fa comparire la spunta cliccabile in Fogli Google, ed è
        # come è fatto il foglio che Lorenzo usa davvero.
        for scarto in range(len(spunte)):
            cella = ws.cell(row=indice, column=5 + scarto)
            cella.value = False
            cella.alignment = centrato
        if riga["link"]:
            cella = ws.cell(row=indice, column=5 + len(spunte))
            cella.hyperlink = riga["link"]
            cella.font = Font(color="FF1155CC", underline="single")

    # --- La formula viva del contatore ------------------------------------
    # [AGGIUNTO 2026-08-05 — task #193] `CONTA.SE` su tutto il blocco delle
    # caselle: il numero cala da solo mentre si spunta, in Excel come in
    # Fogli Google. Un numero fisso scritto qui direbbe sempre la stessa cosa
    # e sarebbe peggio di niente.
    #
    # Il nome inglese `COUNTIF` non e' una svista: dentro il FILE la formula
    # si scrive sempre in inglese, e sono i programmi a mostrarla tradotta a
    # chi la apre in italiano. Scrivendo `CONTA.SE` il foglio uscirebbe con
    # un errore `#NOME?` per tutti.
    if prima_riga_corpo is not None and spunte:
        prima = get_column_letter(5)
        ultima = get_column_letter(4 + len(spunte))
        intervallo = f"{prima}{prima_riga_corpo}:{ultima}{ws.max_row}"
        ws.cell(row=riga_contatore, column=1).value = (
            f'=IF(COUNTIF({intervallo},FALSE)=0,'
            f'"\u2713 Tutto spuntato: si parte.",'
            f'"Mancano ancora "&COUNTIF({intervallo},FALSE)&" caselle da spuntare")'
        )

    larghezze = [13, 21, 14, 58] + [13] * len(spunte) + [42, 50]
    for indice, larghezza in enumerate(larghezze, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = larghezza
    # Il blocco che si congela e quello che si filtra partono dall'intestazione
    # vera: se il pulsante o l'avviso l'hanno spostata in basso, congelare la
    # riga 1 vorrebbe dire tenere fisso il pulsante e far scorrere via i nomi
    # delle colonne, che è il contrario di quello che serve scorrendo 40 voci.
    # La linguetta colorata: si vede prima di aprire il foglio, ed e' il
    # dettaglio piu' economico che dice «questo documento e' di qualcuno».
    ws.sheet_properties.tabColor = identita.excel(identita.NOTTE)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{riga_intestazione + 1}"
    ws.auto_filter.ref = (
        f"A{riga_intestazione}:{get_column_letter(len(intestazioni))}{ws.max_row}"
    )

    # --- Secondo foglio: il programma in breve ----------------------------
    if righe_itinerario:
        wi = wb.create_sheet("Itinerario")
        # Il pulsante serve qui quanto sull'altro foglio, anzi di più: chi
        # guarda "il 12 dove siamo" è esattamente chi vuole riaprire il
        # programma completo. Lo scrive `_scrivi_testata` qui sopra.
        _scrivi_testata(wi, trip, 3)
        if url_itinerario:
            _scrivi_bottone_itinerario(wi, url_itinerario)
        wi.append(["Data", "Luogo", "Programma"])
        intestazione_itinerario = wi.max_row
        for cella in wi[intestazione_itinerario]:
            cella.font = Font(bold=True, color="FFFFFFFF", size=10)
            cella.fill = testata
            cella.alignment = centrato
        wi.row_dimensions[intestazione_itinerario].height = 24
        for riga in righe_itinerario:
            wi.append([riga["data"], riga["luogo"], riga["programma"]])
            for colonna in range(1, 4):
                wi.cell(row=wi.max_row, column=colonna).alignment = a_capo
        for lettera, larghezza in (("A", 22), ("B", 34), ("C", 90)):
            wi.column_dimensions[lettera].width = larghezza
        wi.sheet_properties.tabColor = identita.excel(identita.ORO)
        wi.sheet_view.showGridLines = False
        wi.freeze_panes = f"A{intestazione_itinerario + 1}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
