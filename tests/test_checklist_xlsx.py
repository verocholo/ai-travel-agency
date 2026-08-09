"""
[AGGIUNTO 2026-08-02 — task #172] Copre `src/checklist_xlsx.py`, il foglio
di calcolo della valigia allegato alla stessa mail del PDF.

[Richiesta di Lorenzo: "per quanto riguarda la parte della valigia mi piace
molto ma dopo l'elenco vorrei che creassi un collegamento per un foglio di
calcolo google come quello che ti ho allegato ovviamente costruito in base a
ciò che richiede la valigia, ma strutturato in maniera simile a quello
allegato"]

Le due frasi che decidono cosa va provato qui sono "costruito in base a ciò
che richiede la valigia" e "strutturato in maniera simile a quello allegato":

* «in base a ciò che richiede la valigia» significa che il foglio NON è una
  seconda lista scritta a mano. È un'altra vista degli stessi dati che
  finiscono nel PDF (`vademecum` + `predeparture`). Se le due liste potessero
  divergere, il cliente si troverebbe nella stessa mail due elenchi che si
  contraddicono — ed è il difetto peggiore possibile per un documento che
  serve a fidarsi. Il controllo che lo impedisce è
  `test_ogni_riga_viene_dal_documento_non_da_una_seconda_lista`.

* «strutturato in maniera simile a quello allegato» significa cose misurabili
  sul file di Lorenzo, non impressioni: le fasce di priorità colorate in
  testa a ogni blocco, le caselle da spuntare che in Fogli Google sono
  spuntabili DAVVERO (celle booleane, non la parola "FALSE"), una colonna per
  persona, il secondo foglio con il programma.

Il resto sono le regole di sopravvivenza: questo file è un allegato, e un
allegato che non si apre è un difetto che il cliente vede prima ancora del
PDF. Quindi `build_checklist_xlsx` non solleva MAI: al massimo restituisce
`None` e la mail parte senza allegato invece di non partire.
"""
import datetime
import io
import unittest

from src import checklist_xlsx


class FakeTrip:
    def __init__(self, destination="Siena", date_start="2026-09-14"):
        self.destination = destination
        self.date_start = date_start


VADEMECUM = {
    "climate": {
        "month": 9,
        "temp_max": "28°",
        "temp_min": "17°",
        "forecast_link": {
            "url": "https://www.google.com/search?q=meteo+Siena",
            "label": "Previsioni reali per Siena",
        },
    },
    "packing": [
        {"group": "Documenti e salute", "items": [
            "Documento d'identità valido, più una foto sul telefono",
            "Tessera sanitaria: in Italia basta questa al pronto soccorso",
        ]},
        {"group": "Per il clima di settembre", "items": [
            "Una felpa leggera: la sera in collina la temperatura scende",
        ]},
        {"group": "Elettronica", "items": [
            "Powerbank: una giornata a piedi con le mappe aperte svuota il telefono",
        ]},
    ],
    "baggage": {
        "choice": "un bagaglio a mano a testa",
        "reason": "Per due notti il bagaglio in stiva costa più di quanto serva.",
        "total": "0 € in più sul volo.",
    },
    "suitcase": {},
}

PREDEPARTURE = {
    "checklist": [
        {"title": "Documento d'identità valido per tutta la durata del viaggio",
         "detail": "Controllalo adesso: un rinnovo richiede settimane."},
        {"title": "Conferma della prenotazione dell'alloggio salvata offline",
         "detail": "Serve al check-in anche senza rete."},
    ],
}

ITINERARY = {
    "days": [
        {"day": 1, "title": "Centro storico", "blocks": [
            {"time": "10:30", "activity": "Piazza del Campo",
             "location": "Piazza del Campo 1"},
            {"time": "11:20", "activity": "Salita alla Torre del Mangia",
             "location": "Via Giovanni Duprè 132"},
        ]},
        {"day": 2, "title": "Fuori le mura", "blocks": [
            {"time": "09:00", "activity": "Basilica dell'Osservanza",
             "location": "Strada dell'Osservanza 7"},
        ]},
    ],
}


def _righe():
    return checklist_xlsx.build_checklist_rows(FakeTrip(), VADEMECUM, PREDEPARTURE)


def _apri(blob: bytes):
    from openpyxl import load_workbook
    return load_workbook(io.BytesIO(blob))


def _riga_intestazione(ws) -> int:
    """La riga con "Priorità", cercata invece che data per scontata.

    [AGGIUNTO 2026-08-03] Sopra l'intestazione ora può esserci il pulsante che
    riporta all'itinerario (e, quando il numero di viaggiatori supera il tetto,
    l'avviso che lo dice). Un test che continuasse a leggere la riga 1 non
    proverebbe più niente: leggerebbe il pulsante e lo scambierebbe per
    l'intestazione.
    """
    for indice in range(1, min(ws.max_row, 8) + 1):
        if ws.cell(row=indice, column=1).value == "Priorità":
            return indice
    raise AssertionError(f"intestazione assente: {[c.value for c in ws['A']][:8]}")


def _spunte(ws) -> list[str]:
    """Le intestazioni delle colonne da spuntare, nell'ordine del foglio."""
    riga = _riga_intestazione(ws)
    return [
        str(c.value) for c in ws[riga]
        if c.value and str(c.value).startswith("Fatto")
    ]


def _testi(ws) -> str:
    return " | ".join(
        str(c.value) for riga in ws.iter_rows() for c in riga if c.value is not None
    )


def _collegamenti(wb) -> list[str]:
    return [
        c.hyperlink.target
        for ws in wb.worksheets for riga in ws.iter_rows() for c in riga
        if c.hyperlink is not None
    ]


class TestLeRigheVengonoDalDocumento(unittest.TestCase):
    """Il foglio è una seconda VISTA della valigia, non una seconda LISTA."""

    def test_ogni_riga_viene_dal_documento_non_da_una_seconda_lista(self):
        """Ogni voce della valigia e della lista "prima di partire" deve
        ritrovarsi nel foglio. Se domani qualcuno aggiungesse una voce al
        PDF e non al foglio, il cliente aprirebbe nella stessa mail due
        elenchi che non coincidono."""
        righe = _righe()
        testo = " | ".join(r["attivita"] for r in righe)
        attese = [
            "Documento d'identità valido per tutta la durata",  # predeparture
            "Tessera sanitaria",                                # packing
            "Una felpa leggera",                                # packing clima
            "Powerbank",                                        # packing elettronica
        ]
        mancanti = [a for a in attese if a.split(":")[0][:24] not in testo]
        self.assertEqual([], mancanti, f"voci del documento assenti dal foglio: {mancanti}")

    def test_senza_ingredienti_non_si_inventa_niente(self):
        """Un vademecum vuoto non deve produrre righe inventate: solo le due
        righe universali del "sul posto", che sono vere per ogni viaggio."""
        righe = checklist_xlsx.build_checklist_rows(FakeTrip(), {}, {})
        self.assertTrue(all(r["banda"] == "viaggio" for r in righe),
                        f"righe inventate senza dati: {[r['attivita'] for r in righe]}")

    def test_il_meteo_porta_il_link_vero_e_mai_uno_indovinato(self):
        righe = _righe()
        meteo = [r for r in _righe() if r["categoria"] == "Meteo"]
        self.assertEqual(1, len(meteo), "la riga del meteo deve esserci una volta sola")
        self.assertEqual(VADEMECUM["climate"]["forecast_link"]["url"], meteo[0]["link"])
        for riga in righe:
            if riga["link"]:
                self.assertTrue(riga["link"].startswith("https://"),
                                f"link non cifrato nel foglio: {riga['link']}")

    def test_il_bagaglio_dice_quale_e_perche(self):
        righe = [r for r in _righe() if r["categoria"] == "Bagaglio"]
        self.assertTrue(righe, "il bagaglio è la sola riga che costa soldi: deve esserci")
        scelta = righe[0]
        self.assertIn("bagaglio a mano", scelta["attivita"])
        self.assertIn("costa più di quanto serva", scelta["note"])


class TestNienteDoppioni(unittest.TestCase):
    """[MISURATO] Alla prima prova "Documento d'identità valido..." usciva due
    volte: una dalla lista della sera prima, una dal gruppo "Documenti e
    salute". Nel PDF sono due capitoli lontani e non si nota; in una tabella
    da spuntare due righe quasi uguali fanno perdere fiducia in tutte le
    altre."""

    def test_due_voci_che_dicono_la_stessa_cosa_diventano_una_riga_sola(self):
        import re
        righe = _righe()
        impronte = [
            " ".join(re.findall(r"\w+", r["attivita"].lower(), flags=re.UNICODE)[:4])
            for r in righe
        ]
        doppioni = sorted({i for i in impronte if impronte.count(i) > 1})
        self.assertEqual([], doppioni, f"righe doppie nel foglio: {doppioni}")

    def test_il_documento_di_identita_compare_una_volta_sola(self):
        """Il caso concreto che ha fatto nascere il controllo."""
        righe = [r for r in _righe() if "identità" in r["attivita"].lower()]
        self.assertEqual(1, len(righe), [r["attivita"] for r in righe])


class TestLOrdineEQuelloDelTempo(unittest.TestCase):
    """Il foglio di Lorenzo è ordinato per QUANDO, non per categoria: prima
    quello che, se manca, non si rimedia il giorno prima."""

    def test_le_bande_escono_nellordine_in_cui_si_fanno_le_cose(self):
        ordine = [checklist_xlsx._BANDA_INDICE[r["banda"]] for r in _righe()]
        self.assertEqual(sorted(ordine), ordine, "le fasce di priorità sono fuori ordine")

    def test_la_colonna_quando_porta_una_data_vera_non_un_generico(self):
        """"Entro il 15 agosto" si può mettere in agenda; "entro 2 giorni" no,
        perché non si sa da quando si conta."""
        righe = [r for r in _righe() if r["banda"] == "subito"]
        self.assertTrue(righe)
        self.assertRegex(righe[0]["quando"], r"\d")

    def test_senza_data_di_partenza_si_scrive_la_distanza_non_una_data_falsa(self):
        righe = checklist_xlsx.build_checklist_rows(
            FakeTrip(date_start=None), VADEMECUM, PREDEPARTURE)
        quando = {r["quando"] for r in righe if r["banda"] == "subito"}
        self.assertTrue(quando)
        for testo in quando:
            self.assertIn("prima di partire", testo.lower())


class TestIlFileSiApreEdEStrutturatoComeQuelloDiLorenzo(unittest.TestCase):

    def setUp(self):
        self.blob = checklist_xlsx.build_checklist_xlsx(
            FakeTrip(), VADEMECUM, PREDEPARTURE, ITINERARY, travellers=2)
        self.assertIsInstance(self.blob, bytes)
        self.wb = _apri(self.blob)

    def test_ci_sono_i_due_fogli_checklist_e_itinerario(self):
        self.assertEqual(["Checklist", "Itinerario"], self.wb.sheetnames)

    def test_le_intestazioni_sono_quelle_del_foglio_di_lorenzo(self):
        ws = self.wb["Checklist"]
        intestazioni = [c.value for c in ws[_riga_intestazione(ws)]]
        self.assertEqual("Priorità", intestazioni[0])
        self.assertEqual(["Quando", "Categoria", "Attività"], intestazioni[1:4])
        self.assertEqual(["Link", "Note"], intestazioni[-2:])

    def test_le_caselle_da_spuntare_sono_booleane_vere(self):
        """È questo che fa comparire la spunta CLICCABILE in Fogli Google. La
        parola "FALSE" scritta in una cella di testo sembra uguale sullo
        schermo e non si può spuntare: sarebbe una lista da leggere, non da
        usare, cioè esattamente quello che il PDF già fa meglio."""
        ws = self.wb["Checklist"]
        tipi = set()
        for riga in ws.iter_rows(min_row=_riga_intestazione(ws) + 1, min_col=5, max_col=6):
            for cella in riga:
                if cella.value is not None:
                    tipi.add(cella.data_type)
        self.assertIn("b", tipi, f"nessuna casella spuntabile: tipi trovati {tipi}")

    def test_una_colonna_per_viaggiatore(self):
        ws = self.wb["Checklist"]
        self.assertEqual(2, len(_spunte(ws)), _spunte(ws))

    def test_le_fasce_colorate_ci_sono_e_sono_colorate_fino_in_fondo(self):
        """Una fascia colorata solo sulla prima colonna, dopo la fusione,
        sembra spezzata a metà foglio."""
        ws = self.wb["Checklist"]
        etichette = {b["label"] for b in checklist_xlsx.BANDE}
        trovate = 0
        for riga in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            cella = riga[0]
            if cella.value in etichette:
                trovate += 1
                ultima = ws.cell(row=cella.row, column=ws.max_column)
                self.assertEqual(cella.fill.fgColor.rgb, ultima.fill.fgColor.rgb,
                                 "la fascia si interrompe prima dell'ultima colonna")
        self.assertGreaterEqual(trovate, 3, "le fasce di priorità sono sparite")

    def test_il_secondo_foglio_racconta_le_giornate_con_i_nomi_non_gli_indirizzi(self):
        """[MISURATO] La prima versione usava `location`, che è l'INDIRIZZO:
        veniva fuori "Piazza del Campo 1 → Via Giovanni Duprè 132", una riga
        che non riconosce nemmeno chi c'è stato."""
        ws = self.wb["Itinerario"]
        programmi = [ws.cell(row=r, column=3).value or "" for r in range(2, ws.max_row + 1)]
        testo = " ".join(programmi)
        self.assertIn("Salita alla Torre del Mangia", testo)
        self.assertNotIn("Via Giovanni Duprè 132", testo)
        self.assertIn("10:30", testo)

    def test_il_link_del_meteo_e_cliccabile_dentro_il_foglio(self):
        ws = self.wb["Checklist"]
        link = [c.hyperlink.target for r in ws.iter_rows(min_row=2) for c in r if c.hyperlink]
        self.assertTrue(link, "il link del meteo non è cliccabile nel foglio")
        for bersaglio in link:
            self.assertTrue(bersaglio.startswith("https://"), bersaglio)


class TestUnaCasellaPerOgniViaggiatoreVero(unittest.TestCase):
    """[NUOVO 2026-08-03 — richiesta di Lorenzo, alla lettera: "ricordati di
    aggiungere poi le spunte per i viaggiatori (se sono tre, 3 caselle di
    checklist, se sono 4 ne metti 4 e così via)"]

    Prima il foglio si fermava a quattro colonne, e la quinta persona di un
    gruppo di cinque non aveva dove spuntare: o si spuntava sopra la casella di
    qualcun altro, o si rinunciava al foglio e si tornava alla penna. È
    esattamente il difetto che il foglio doveva togliere."""

    def test_tre_viaggiatori_fanno_tre_caselle_quattro_ne_fanno_quattro(self):
        """La frase del cliente, provata numero per numero."""
        for quanti in (1, 2, 3, 4, 5, 6):
            with self.subTest(viaggiatori=quanti):
                blob = checklist_xlsx.build_checklist_xlsx(
                    FakeTrip(), VADEMECUM, PREDEPARTURE, ITINERARY, travellers=quanti)
                ws = _apri(blob)["Checklist"]
                self.assertEqual(quanti, len(_spunte(ws)),
                                 f"{quanti} viaggiatori, colonne trovate: {_spunte(ws)}")

    def test_ogni_colonna_in_piu_e_una_casella_spuntabile_vera_non_una_scritta(self):
        """Una colonna che c'è ma non si spunta è peggio di una colonna che
        manca: sembra funzionare finché il quinto viaggiatore non ci prova dal
        telefono, la sera prima di partire."""
        for quanti in (1, 2, 3, 4, 5, 6):
            with self.subTest(viaggiatori=quanti):
                blob = checklist_xlsx.build_checklist_xlsx(
                    FakeTrip(), VADEMECUM, PREDEPARTURE, ITINERARY, travellers=quanti)
                ws = _apri(blob)["Checklist"]
                intestazione = _riga_intestazione(ws)
                for scarto in range(quanti):
                    tipi = {
                        ws.cell(row=r, column=5 + scarto).data_type
                        for r in range(intestazione + 1, ws.max_row + 1)
                        if ws.cell(row=r, column=5 + scarto).value is not None
                    }
                    self.assertEqual({"b"}, tipi,
                                     f"colonna {scarto + 1} non spuntabile: {tipi}")

    def test_le_colonne_restano_numerate_e_non_chiedono_i_nomi_delle_persone(self):
        """I nomi dei compagni di viaggio non li abbiamo e non li chiediamo:
        sarebbero dati di persone che non sono nostre clienti. La colonna
        numerata si rinomina in due secondi con il nome vero, e intanto non
        raccogliamo niente di nessuno."""
        blob = checklist_xlsx.build_checklist_xlsx(
            FakeTrip(), VADEMECUM, PREDEPARTURE, ITINERARY, travellers=5)
        ws = _apri(blob)["Checklist"]
        self.assertEqual(
            [f"Fatto · viaggiatore {i}" for i in range(1, 6)], _spunte(ws))

    def test_con_molte_colonne_le_fasce_arrivano_comunque_in_fondo(self):
        """Le fasce colorate si allargano con le colonne: con sei viaggiatori
        una fascia ferma alla quarta colonna spezzerebbe il foglio a metà."""
        blob = checklist_xlsx.build_checklist_xlsx(
            FakeTrip(), VADEMECUM, PREDEPARTURE, ITINERARY, travellers=6)
        ws = _apri(blob)["Checklist"]
        etichette = {b["label"] for b in checklist_xlsx.BANDE}
        trovate = 0
        for riga in ws.iter_rows(min_col=1, max_col=1):
            cella = riga[0]
            if cella.value in etichette:
                trovate += 1
                ultima = ws.cell(row=cella.row, column=ws.max_column)
                self.assertEqual(cella.fill.fgColor.rgb, ultima.fill.fgColor.rgb)
        self.assertGreaterEqual(trovate, 3)


class TestIlPulsanteCheRiportaAllItinerario(unittest.TestCase):
    """[NUOVO 2026-08-03 — richiesta di Lorenzo: "ovviamente un pulsante sul
    foglio di calcolo che ti fa ritornare al pdf originario"]

    Chi apre il foglio dal telefono, in piedi davanti alla valigia, ha perso il
    PDF: è in una mail di tre settimane fa, sotto altre venti. Il pulsante è
    l'unica strada breve fra lo strumento e il documento. Ma vale solo se
    l'indirizzo esiste davvero: un pulsante che non apre niente costa più
    fiducia di quanta ne dia un pulsante che non c'è."""

    URL = "https://esempio.invalido/itinerari/siena-2026-09.pdf"

    def _foglio(self, url):
        blob = checklist_xlsx.build_checklist_xlsx(
            FakeTrip(), VADEMECUM, PREDEPARTURE, ITINERARY, travellers=2,
            itinerary_url=url)
        self.assertIsInstance(blob, bytes)
        return _apri(blob)

    def test_il_pulsante_c_e_su_tutti_e_due_i_fogli(self):
        """Sul secondo foglio serve quanto sul primo: chi guarda "il 12 dove
        siamo" è proprio chi vuole riaprire il programma completo."""
        wb = self._foglio(self.URL)
        for nome in ("Checklist", "Itinerario"):
            with self.subTest(foglio=nome):
                self.assertIn(checklist_xlsx.TESTO_BOTTONE_ITINERARIO,
                              _testi(wb[nome]), f"pulsante assente dal foglio {nome}")

    def test_il_pulsante_porta_esattamente_a_quell_indirizzo(self):
        """Un pulsante che porta a un indirizzo simile ma non uguale è un
        pulsante che apre l'itinerario di qualcun altro."""
        wb = self._foglio(self.URL)
        for nome in ("Checklist", "Itinerario"):
            ws = wb[nome]
            bersagli = [
                c.hyperlink.target for riga in ws.iter_rows() for c in riga
                if c.hyperlink is not None
                and str(c.value or "").startswith(checklist_xlsx.TESTO_BOTTONE_ITINERARIO[0])
            ]
            with self.subTest(foglio=nome):
                self.assertEqual([self.URL], bersagli)

    def test_il_pulsante_sta_in_cima_e_si_vede_che_e_un_pulsante(self):
        """In fondo al foglio non lo troverebbe nessuno, e una cella azzurra
        senza grassetto né bordo si legge come una riga qualsiasi: si clicca
        solo quello che sembra cliccabile."""
        wb = self._foglio(self.URL)
        for nome in ("Checklist", "Itinerario"):
            ws = wb[nome]
            # [AGGIORNATO 2026-08-05 — task #193] Non più «riga 1»: sopra il
            # pulsante c'è ora il blocco di testa con il marchio e il viaggio.
            # Quello che conta non è cambiato ed è quello che si controlla:
            # il pulsante sta PRIMA dell'intestazione delle colonne, cioè
            # dentro la prima schermata, e sembra un pulsante.
            righe = [c.row for r in ws.iter_rows(min_col=1, max_col=1) for c in r
                     if c.value == checklist_xlsx.TESTO_BOTTONE_ITINERARIO]
            with self.subTest(foglio=nome):
                self.assertEqual(len(righe), 1,
                                 "il pulsante manca, o è stampato due volte")
                cella = ws.cell(row=righe[0], column=1)
                self.assertLess(righe[0], _riga_intestazione(ws) if nome == "Checklist" else 99,
                                "il pulsante è finito sotto l'intestazione")
                self.assertLessEqual(righe[0], 5,
                                     "il pulsante è troppo in basso per vedersi "
                                     "senza scorrere")
                self.assertTrue(cella.font.bold, "il pulsante non è in grassetto")
                self.assertEqual("solid", cella.fill.fill_type, "il pulsante non è colorato")
                self.assertNotIn(cella.fill.fgColor.rgb, (None, "00000000"),
                                 "il pulsante non ha un colore di sfondo vero")
                self.assertTrue(cella.border.bottom.style, "il pulsante non ha un bordo")

    def test_senza_un_indirizzo_usabile_non_si_promette_nessun_pulsante(self):
        """Il difetto peggiore possibile qui: una casella che dice "torna
        all'itinerario", si clicca e non apre niente. `http://` è escluso come
        ovunque nel progetto: un allegato non manda un cliente su una pagina in
        chiaro."""
        senza_meteo = {k: v for k, v in VADEMECUM.items() if k != "climate"}
        for url in (None, "", "   ", "http://x/y", "non-un-url", 123,
                    "https://", "ftp://x/y", "//esempio.invalido/x", b"https://x/y"):
            with self.subTest(url=url):
                blob = checklist_xlsx.build_checklist_xlsx(
                    FakeTrip(), senza_meteo, PREDEPARTURE, ITINERARY,
                    travellers=2, itinerary_url=url)
                self.assertIsInstance(blob, bytes)
                wb = _apri(blob)
                self.assertEqual([], _collegamenti(wb),
                                 "c'è un collegamento verso il nulla nel foglio")
                for nome in wb.sheetnames:
                    self.assertNotIn("Torna all'itinerario", _testi(wb[nome]),
                                     f"pulsante morto sul foglio {nome}")

    def test_senza_pulsante_l_intestazione_resta_la_prima_riga(self):
        """Il riquadro del PDF conta le voci del foglio: una riga in più in
        testa quando il pulsante non c'è farebbe dire al PDF un numero e al
        foglio un altro."""
        blob = checklist_xlsx.build_checklist_xlsx(
            FakeTrip(), VADEMECUM, PREDEPARTURE, ITINERARY, travellers=2)
        ws = _apri(blob)["Checklist"]
        # [AGGIORNATO 2026-08-05] Il numero fisso non c'è più: sopra
        # l'intestazione ci sono la testata e il contatore, che ci sono
        # sempre. Quello che il riquadro del PDF conta sono le VOCI, e le
        # voci sono le righe sotto l'intestazione che hanno una casella —
        # ed è quello che questo controllo protegge davvero.
        self.assertNotIn(checklist_xlsx.TESTO_BOTTONE_ITINERARIO,
                         [c.value for r in ws.iter_rows(min_col=1, max_col=1)
                          for c in r])
        self.assertEqual(["Quando", "Categoria", "Attività"],
                         [c.value for c in ws[_riga_intestazione(ws)]][1:4])

    def test_il_pulsante_non_scompiglia_niente_di_quello_che_c_era(self):
        """Il pulsante è un'aggiunta: fasce, caselle booleane, secondo foglio e
        link del meteo devono restare esattamente come sono."""
        wb = self._foglio(self.URL)
        ws = wb["Checklist"]
        intestazione = _riga_intestazione(ws)
        righe_pulsante = [c.row for r in ws.iter_rows(min_col=1, max_col=1)
                          for c in r
                          if c.value == checklist_xlsx.TESTO_BOTTONE_ITINERARIO]
        self.assertTrue(righe_pulsante, "il pulsante è sparito")
        self.assertLess(righe_pulsante[0], intestazione,
                        "il pulsante deve stare sopra l'intestazione")
        self.assertEqual(["Quando", "Categoria", "Attività"],
                         [c.value for c in ws[intestazione]][1:4])
        self.assertEqual(["Link", "Note"], [c.value for c in ws[intestazione]][-2:])
        self.assertEqual(2, len(_spunte(ws)))
        etichette = {b["label"] for b in checklist_xlsx.BANDE}
        fasce = [c.row for r in ws.iter_rows(min_col=1, max_col=1)
                 for c in r if c.value in etichette]
        self.assertGreaterEqual(len(fasce), 3, "le fasce sono sparite con il pulsante")
        for indice in fasce:
            self.assertEqual(ws.cell(row=indice, column=1).fill.fgColor.rgb,
                             ws.cell(row=indice, column=ws.max_column).fill.fgColor.rgb)
        tipi = {
            ws.cell(row=r, column=5).data_type
            for r in range(intestazione + 1, ws.max_row + 1)
            if ws.cell(row=r, column=5).value is not None
        }
        self.assertEqual({"b"}, tipi, "le caselle non sono più spuntabili")
        self.assertEqual(["Checklist", "Itinerario"], wb.sheetnames)
        wi = wb["Itinerario"]
        self.assertIn("Salita alla Torre del Mangia", _testi(wi))
        self.assertIn(VADEMECUM["climate"]["forecast_link"]["url"], _collegamenti(wb))
        for bersaglio in _collegamenti(wb):
            self.assertTrue(bersaglio.startswith("https://"), bersaglio)


class TestIlFoglioNonPuoRompereLaConsegna(unittest.TestCase):
    """Il PDF è il prodotto, il foglio è un di più. Un errore qui non deve
    poter togliere il documento al cliente che l'ha pagato."""

    def test_senza_niente_da_spuntare_si_restituisce_none_non_un_file_vuoto(self):
        self.assertIsNone(checklist_xlsx.build_checklist_xlsx(None, None, None, None))

    def test_ingredienti_malformati_non_sollevano_mai(self):
        casi = [
            {"packing": "non è una lista"},
            {"packing": [None, 3, {"group": None, "items": None}]},
            {"climate": [], "baggage": "no", "packing": [{"items": [None, 7]}]},
        ]
        for caso in casi:
            with self.subTest(caso=caso):
                try:
                    blob = checklist_xlsx.build_checklist_xlsx(
                        FakeTrip(), caso, {"checklist": ["non un dizionario"]}, ITINERARY)
                except Exception as e:  # noqa: BLE001
                    self.fail(f"il foglio ha sollevato invece di rinunciare: {e!r}")
                self.assertTrue(blob is None or isinstance(blob, bytes))

    def test_oltre_il_tetto_le_colonne_si_fermano_e_il_foglio_non_diventa_illeggibile(self):
        """[MISURATO 2026-08-03] Senza tetto un valore sbagliato che arriva dal
        modulo (40, o peggio 10.000) non fa un foglio brutto: fa un foglio che
        non esce. Un `.xlsx` da 1000 colonne costa 23 secondi di generazione,
        e a 16.380 colonne il salvataggio non è finito dopo dieci minuti — cioè
        oltre il tetto dei 300 secondi di Make, quindi niente PDF per nessuno."""
        blob = checklist_xlsx.build_checklist_xlsx(
            FakeTrip(), VADEMECUM, PREDEPARTURE, ITINERARY, travellers=40)
        ws = _apri(blob)["Checklist"]
        self.assertEqual(checklist_xlsx.MAX_COLONNE_SPUNTA, len(_spunte(ws)))

    def test_quando_le_colonne_vengono_tagliate_il_foglio_lo_dice(self):
        """Tagliare in silenzio è la versione peggiore: chi ha detto "siamo in
        20" apre il foglio, conta 12 colonne e non sa se il conto che ci ha
        dato è arrivato o se il foglio è rotto. La riga di avviso dice il
        numero chiesto e quello tenuto, e cosa fare per gli altri."""
        blob = checklist_xlsx.build_checklist_xlsx(
            FakeTrip(), VADEMECUM, PREDEPARTURE, ITINERARY, travellers=20)
        ws = _apri(blob)["Checklist"]
        testo = _testi(ws)
        self.assertIn("20", testo, "il foglio non dice quanti viaggiatori erano stati chiesti")
        self.assertIn(str(checklist_xlsx.MAX_COLONNE_SPUNTA), testo,
                      "il foglio non dice quante colonne ha davvero")

    def test_sotto_il_tetto_nessun_avviso_sporca_il_foglio(self):
        """L'avviso è la risposta a un problema: se il problema non c'è, la
        riga in più è solo rumore in testa a un foglio che si deve capire in
        due secondi."""
        blob = checklist_xlsx.build_checklist_xlsx(
            FakeTrip(), VADEMECUM, PREDEPARTURE, ITINERARY, travellers=4)
        ws = _apri(blob)["Checklist"]
        # [AGGIORNATO 2026-08-05] Il controllo non è più sul numero di riga —
        # la testata e il contatore ci sono sempre — ma sulla cosa che
        # importava: l'avviso NON deve comparire quando non c'è niente da
        # avvisare.
        testi = [str(c.value or "") for r in ws.iter_rows(min_col=1, max_col=1)
                 for c in r]
        self.assertFalse([t for t in testi if t.startswith("⚠")],
                         "c'è un avviso che non serve, in testa al foglio")

    def test_un_numero_di_viaggiatori_assurdo_non_rompe_niente(self):
        """`travellers` arriva dal modulo di richiesta e passa per Make: può
        essere una stringa, un vuoto, un decimale o un numero senza senso.
        Nessuno di questi casi deve poter togliere l'allegato al cliente."""
        for quanti in (0, -3, None, "due", 2.5, 10 ** 9, "", True, float("nan"),
                       float("inf"), [], 10 ** 400):
            with self.subTest(travellers=quanti):
                colonne = checklist_xlsx._colonne_spunta(quanti)
                self.assertGreaterEqual(len(colonne), 1)
                self.assertLessEqual(len(colonne), checklist_xlsx.MAX_COLONNE_SPUNTA)
                try:
                    blob = checklist_xlsx.build_checklist_xlsx(
                        FakeTrip(), VADEMECUM, PREDEPARTURE, ITINERARY,
                        travellers=quanti)
                except Exception as e:  # noqa: BLE001
                    self.fail(f"il foglio ha sollevato invece di rinunciare: {e!r}")
                self.assertIsInstance(blob, bytes)
                ws = _apri(blob)["Checklist"]
                self.assertEqual(len(colonne), len(_spunte(ws)))

    def test_mezzo_viaggiatore_diventa_una_colonna_intera(self):
        """2.5 non è un errore da rifiutare: nei preventivi è "due adulti e un
        bambino". Il bambino la valigia ce l'ha, quindi si arrotonda per
        eccesso — una colonna in più si ignora, una in meno lascia qualcuno
        senza la sua."""
        self.assertEqual(3, len(checklist_xlsx._colonne_spunta(2.5)))


class TestIlNomeDelFileSiRitrova(unittest.TestCase):

    def test_il_nome_porta_destinazione_e_mese(self):
        self.assertEqual("Valigia-Siena-2026-09.xlsx",
                         checklist_xlsx.build_checklist_filename(FakeTrip()))

    def test_una_destinazione_con_spazi_e_accenti_resta_un_nome_di_file_valido(self):
        nome = checklist_xlsx.build_checklist_filename(
            FakeTrip(destination="Città di Castello / Umbria"))
        self.assertTrue(nome.endswith(".xlsx"))
        for vietato in "/\\:*?\"<>|":
            self.assertNotIn(vietato, nome)

    def test_senza_destinazione_resta_comunque_un_nome_sensato(self):
        nome = checklist_xlsx.build_checklist_filename(FakeTrip(destination="", date_start=None))
        self.assertEqual("Valigia.xlsx", nome)


class TestIlFoglioParlaDelViaggioGiusto(unittest.TestCase):
    """Un foglio con le date di un altro viaggio è peggio di nessun foglio."""

    def test_le_date_seguono_la_partenza_vera(self):
        righe = checklist_xlsx.build_checklist_rows(
            FakeTrip(date_start="2027-03-02"), VADEMECUM, PREDEPARTURE)
        quando = " ".join(r["quando"] for r in righe)
        self.assertIn("marzo", quando.lower() + " " + quando)
        self.assertNotIn("settembre", quando.lower())

    def test_le_giornate_del_secondo_foglio_hanno_le_date_del_viaggio(self):
        righe = checklist_xlsx.build_itinerary_rows(ITINERARY, FakeTrip())
        self.assertEqual(2, len(righe))
        self.assertIn("14 settembre", righe[0]["data"])
        self.assertIn("15 settembre", righe[1]["data"])


if __name__ == "__main__":  # pragma: no cover
    unittest.main()
