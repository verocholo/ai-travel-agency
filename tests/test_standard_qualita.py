"""
LO STANDARD — i controlli che rendono superfluo ripetere la richiesta.

[CREATO 2026-08-02 — task #169, richiesta esplicita di Lorenzo:
"fai in modo che lavorare così sia il lavoro standard senza bisogno ogni
volta dello stesso comando, mantieni sempre questa qualità con tutte le
caratteristiche"]

Perché questo file esiste separato dagli altri
----------------------------------------------
Gli altri file di test verificano che un pezzo di codice faccia quello che
dice. Questo verifica una cosa diversa: che il DOCUMENTO CONSEGNATO AL
CLIENTE continui ad avere tutte le caratteristiche che Lorenzo ha chiesto,
una richiesta alla volta, nell'arco di settimane. Sono richieste che il
codice può soddisfare oggi e perdere domani senza che nessun test unitario
se ne accorga — perché nessuna di esse appartiene a un modulo solo.

Ogni controllo qui sotto porta scritto QUALE richiesta protegge, con le
parole di Lorenzo. Se uno di questi fallisce, non è "un test rotto": è una
caratteristica del prodotto che è appena sparita.

Su cosa girano i controlli
--------------------------
Sul campione di `scripts_sample_pdf.py`, cioè lo STESSO documento che
Lorenzo apre per giudicare il lavoro. È una scelta deliberata: se domani
qualcuno impoverisse il campione per far passare un controllo, i controlli
cadrebbero insieme al campione invece di continuare a dire "verde" su una
vetrina svuotata. Un esempio che si aggiusta da sé non è un esempio.

Cosa gira sull'HTML e cosa sulla carta
--------------------------------------
Quasi tutto gira sull'HTML renderizzato, che è il documento vero prima
dell'impaginazione: così i controlli funzionano ovunque, anche dove
`wkhtmltopdf` non è installato.

Un difetto però sull'HTML non si vede per definizione: le PAGINE MEZZE
VUOTE. Quante schede entrino in un foglio lo decide l'altezza che prendono
una volta stampate, e quella la sa solo il motore di stampa. Per quel solo
controllo (`TestLaCartaNonRestaMezzaVuota`) il campione viene convertito
davvero, e se gli strumenti mancano il controllo si salta invece di
mentire.
"""
import html as _html
import re
import pathlib
import shutil
import subprocess
import tempfile
import unittest

import scripts_sample_pdf
from src import pdf_renderer


def _documento() -> str:
    """L'HTML del campione completo, montato una volta sola."""
    if not hasattr(_documento, "_cache"):
        itinerary, trip, kwargs, errori = scripts_sample_pdf.build_sample_render_kwargs()
        assert not errori, f"il campione monta con sezioni cadute: {errori}"
        _documento._cache = pdf_renderer.render_html(itinerary, trip, **kwargs)
    return _documento._cache


def _leggibile(frammento: str) -> str:
    """Il testo come lo LEGGE il cliente, non come lo scrive il renderer.

    [AGGIUNTO 2026-08-02 (quater)] Cercare "Architect's Tips" dentro l'HTML
    grezzo non trova niente: nel documento c'è `Architect&#x27;s Tips`, perché
    l'apostrofo viene escapato — e i titoli dei capitoli di questo prodotto
    sono pieni di apostrofi ("a colpo d'occhio", "com'è andata"). Un controllo
    che cerca la forma scritta invece di quella letta fallisce su un dettaglio
    di codifica e fa credere che manchi un capitolo che c'è. Peggio: lo stesso
    errore al contrario passerebbe in silenzio.
    """
    return _html.unescape(frammento)


class TestNienteVaPerso(unittest.TestCase):
    """Ogni capitolo che Lorenzo ha chiesto, una richiesta alla volta, deve
    esserci nel documento consegnato. Il modo tipico in cui una sezione
    sparisce non è che qualcuno la cancelli: è che un errore dentro il suo
    costruttore la faccia cadere in silenzio (vedi il `try/except` con
    `_record()` in `pdf_extras.build_pdf_sections`) e il documento esca
    ugualmente, più povero, senza che nessuno se ne accorga."""

    # (frammento nel corpo, frammento nell'indice, richiesta che protegge)
    #
    # I due frammenti sono distinti di proposito: l'indice di copertina usa i
    # titoli lunghi ("La selezione: dove mangiare, cosa fare") mentre il corpo
    # separa le sezioni con titoli corti ("Dove mangiare", "Cosa fare"). Un
    # controllo che cercasse la stessa stringa in tutti e due i posti si
    # potrebbe soddisfare solo impoverendo uno dei due.
    CAPITOLI = [
        ("Il tuo viaggio, a colpo d'occhio", "a colpo d'occhio", "sintesi d'apertura"),
        ("Il tuo alloggio", "Il tuo alloggio", "alloggio con prezzo reale"),
        ("Dove mangiare", "dove mangiare", "ristoranti"),
        ("Cosa fare", "cosa fare", "intrattenimenti in funzione del tipo di vacanza"),
        ("Il programma, giorno per giorno", "Il programma, giorno per giorno", "il cuore del prodotto"),
        ("Stima dei costi", "Stima dei costi", "\"segnare ogni costo\""),
        ("Prima di partire", "Prima di partire", "controllo di freschezza"),
        ("Vademecum", "Vademecum", "\"vademecum di viaggio\" e valigia"),
        ("Architect's Tips", "Architect's Tips", "consigli dell'Architetto per direttrici"),
        ("se piove", "Piani B", "\"piani B se piove\""),
        ("Guide turistiche tascabili", "Guide turistiche tascabili", "\"una guida per ogni cosa che lo richieda\""),
        # "Facci sapere com'è andata" NON sta in questa lista: è l'unico
        # capitolo condizionato, perché è l'unico che dipende da una
        # configurazione esterna (il modulo Tally). Ha una classe di
        # controlli tutta sua, qui sotto, che verifica ENTRAMBI i modi —
        # altrimenti toglierlo da qui sarebbe solo un modo di far tacere
        # un controllo scomodo.
    ]

    def test_tutti_i_capitoli_chiesti_sono_nel_documento(self):
        testo = _leggibile(_documento())
        mancanti = [
            f"{frammento!r} ({motivo})"
            for frammento, _, motivo in self.CAPITOLI
            if frammento not in testo
        ]
        self.assertEqual([], mancanti, f"capitoli spariti dal documento: {mancanti}")

    def test_tutti_i_capitoli_sono_anche_nellindice_di_copertina(self):
        """Un capitolo che c'è ma non è indicizzato è un capitolo che il
        cliente non trova. L'indice sta in copertina: si controlla lì."""
        html = _documento()
        copertina = html.split("class='cover", 1)[1].split("class='header'", 1)[0]
        indice = _leggibile(copertina.split("Cosa troverai dentro", 1)[1]).lower()
        mancanti = [
            f"{voce!r} ({motivo})"
            for _, voce, motivo in self.CAPITOLI
            if voce.lower() not in indice
        ]
        self.assertEqual([], mancanti, f"capitoli assenti dall'indice: {mancanti}")


class TestICollegamentiPortanoDaQualchePartE(unittest.TestCase):
    """[Richiesta del 2026-08-02: "i collegamenti non funzionano: quello per
    la guida turistica che porta in fondo al documento non funziona... non
    funziona nemmeno il collegamento per le recensioni in fondo al
    documento."] Un collegamento morto è peggio di un collegamento assente:
    promette e non mantiene, e il cliente lo scopre solo dopo aver
    cliccato."""

    def test_nessun_collegamento_interno_punta_nel_vuoto(self):
        html = _documento()
        bersagli = set(re.findall(r"id='([^']+)' class='anchor-probe'", html))
        partenze = set(re.findall(r"href='#([^']+)'", html))
        self.assertTrue(partenze, "un documento senza collegamenti interni non è questo prodotto")
        self.assertEqual(
            set(), partenze - bersagli,
            f"collegamenti interni senza bersaglio: {sorted(partenze - bersagli)}",
        )

    def test_ogni_scheda_di_guida_e_raggiungibile_dal_programma(self):
        """La guida in fondo al documento serve a poco se dal programma non
        ci si arriva: è esattamente il difetto segnalato."""
        html = _documento()
        schede = set(re.findall(r"id='(guida-[^']+)' class='anchor-probe'", html))
        self.assertTrue(schede, "il campione deve avere guide tascabili")
        orfane = [a for a in schede if f"href='#{a}'" not in html]
        self.assertEqual([], orfane, f"guide non raggiungibili dal programma: {orfane}")

    def test_il_collegamento_alla_recensione_e_coerente_col_capitolo(self):
        """[AGGIORNATO 2026-08-03] Prima questo pretendeva l'àncora sempre.
        Ora àncora, voce d'indice e capitolo o ci sono tutti e tre o non
        c'è nessuno dei tre: è quella la proprietà che protegge il
        cliente, non la presenza incondizionata di un'àncora."""
        html = _documento()
        ancora = "id='recensione' class='anchor-probe'" in html
        rimando = "href='#recensione'" in html
        capitolo = "Facci sapere com&#x27;è andata" in html or "Facci sapere com'è andata" in html
        self.assertEqual(
            {ancora, rimando, capitolo}, {ancora},
            f"recensione incoerente: àncora={ancora} indice={rimando} capitolo={capitolo}",
        )

    def test_nessun_indirizzo_web_in_chiaro_non_cifrato(self):
        """Un link non cifrato dentro un documento che il cliente apre sul
        telefono è un difetto di sicurezza, non di stile."""
        self.assertNotIn("http://", _documento())


class TestIlCapitoloRecensioneSegueIlModulo(unittest.TestCase):
    """[Richiesta del 2026-08-03: "il link di tally non funziona ancora"]

    Il capitolo della recensione è l'unico del documento che dipende da
    una cosa che non sta nel codice: la URL del modulo Tally, che vive
    nella variabile d'ambiente `FEEDBACK_FORM_URL` su Render. Ha quindi
    DUE modi leciti, e il controllo deve dire quale dei due è attivo
    invece di pretenderne uno solo:

    * modulo configurato → capitolo, voce d'indice e riquadro "Rispondi
      qui" con la URL vera;
    * modulo assente → niente capitolo, niente voce d'indice, niente
      àncora. Non una versione ridotta: proprio niente.

    Il modo sbagliato — quello che ha prodotto la lamentela di Lorenzo —
    è il terzo: il capitolo che esce lo stesso, fa due domande al
    cliente e non gli dà nessun posto in cui rispondere. Da fuori è
    indistinguibile da un link rotto.
    """

    @staticmethod
    def _modulo_configurato() -> bool:
        _, _, kwargs, _ = scripts_sample_pdf.build_sample_render_kwargs()
        return bool((kwargs.get("feedback_link") or {}).get("url"))

    def test_i_due_modi_sono_entrambi_coerenti(self):
        html = _leggibile(_documento())
        if self._modulo_configurato():
            self.assertIn("Facci sapere com'è andata", html)
            self.assertIn("Rispondi qui", html)
        else:
            self.assertNotIn("Facci sapere com'è andata", html)
            self.assertNotIn("Rispondi qui", html)

    def test_il_riquadro_non_promette_mai_un_posto_che_non_ce(self):
        """L'invariante che vale in tutti e due i modi, ed è quella che
        conta: non esiste un documento che dica "Rispondi qui" senza
        avere accanto un indirizzo https a cui rispondere."""
        html = _documento()
        if "Rispondi qui" in html:
            coda = html.split("Rispondi qui", 1)[1][:600]
            self.assertIn("https://", coda,
                          "'Rispondi qui' senza nessun indirizzo a cui rispondere")

    def test_nessun_modulo_di_esempio_finisce_nel_campione(self):
        """La causa prima della lamentela: nel campione c'era
        `https://tally.so/r/ESEMPIO`, che è un 404. Un indirizzo
        inventato dentro la vetrina è peggio di nessun indirizzo."""
        html = _documento()
        for finto in ("tally.so/r/ESEMPIO", "tally.so/r/esempio", "example.com", "localhost"):
            self.assertNotIn(finto, html)


class TestLeCartineCiSono(unittest.TestCase):
    """[Richiesta del 2026-08-02: "le cartine: le hai completamente rimosse,
    attieniti a ciò che ti avevo detto in precedenza per queste"] Le cartine
    erano già state chieste, tolte da una regressione e richieste di nuovo.
    Questo è il controllo che impedisce il terzo giro."""

    def test_ogni_giornata_ha_la_sua_cartina(self):
        itinerary, _, _, _ = scripts_sample_pdf.build_sample_render_kwargs()
        html = _documento()
        giornate = len(itinerary["days"])
        cartine = html.count("class='day-map-img'") or html.count("class='day-map'")
        self.assertGreaterEqual(
            cartine, giornate,
            f"{giornate} giornate ma solo {cartine} cartine: qualcuna è sparita",
        )

    def test_ogni_cartina_ha_la_sua_legenda(self):
        """Una cartina con marker numerati e senza legenda costringe a
        indovinare: era il difetto visto sul PDF vero."""
        html = _documento()
        self.assertIn("legenda", html.lower())

    def test_le_cartine_dichiarano_di_essere_schemi_non_percorsi_di_guida(self):
        """Onestà sui limiti: le linee sono rette fra coordinate reali, non
        un percorso calcolato. Se il documento non lo dice, mente."""
        html = _documento()
        self.assertIn("indicative", html.lower())


class TestLaCartinaNonEQuellaDellaCitta(unittest.TestCase):
    """[Richiesta del 2026-08-02, con la foto alla mano: "ora quella parte è
    fatta bene ma manca la cartina, ci sono solamente i vettori ma la cartina
    in sé manca"]

    Le cartine sono già state chieste due volte e sono già tornate indietro
    due volte, ogni volta per un motivo diverso: prima non c'erano affatto,
    poi c'erano i pallini ma non le strade sotto. La correzione di oggi non
    sceglie più fra le due sorgenti — chiede a Google solo lo SFONDO e ci
    disegna sopra i nostri pallini — e questi controlli bloccano i due modi
    in cui quel meccanismo può guastarsi in silenzio:

      1. si consegna lo sfondo NUDO, cioè la cartina della città invece che
         quella della giornata: bella, con le strade, e senza una sola tappa;
      2. si torna allo schema senza dirlo, cioè la regressione già vista.

    Un difetto che il cliente vede in una foto e noi no è esattamente il tipo
    di difetto che va fermato da un controllo, non da una rilettura."""

    @staticmethod
    def _sfondo(width=1280, height=876) -> bytes:
        """Uno sfondo stradale finto: serve solo a essere RICONOSCIBILE dopo,
        cioè a poter dire "questo è ancora lo sfondo, nessuno ci ha disegnato
        sopra". Per questo è a tinta unita."""
        import io

        from PIL import Image

        buf = io.BytesIO()
        Image.new("RGB", (width, height), (233, 231, 227)).save(buf, format="PNG")
        return buf.getvalue()

    _BASE_MAP = {"center": (43.3181, 11.3307), "zoom": 17,
                 "size": (640, 438), "scale": 2}

    @staticmethod
    def _plan(png=None, base_map=None):
        plan = {
            "day": 1,
            "title": "Centro storico",
            "hotel_point": (43.3167, 11.3300),
            "hotel_name": "Palazzo Ravizza",
            "stops": [
                {"label": "1", "name": "Piazza del Campo",
                 "point": (43.3182, 11.3315), "color": "blue"},
                {"label": "2", "name": "Duomo di Siena",
                 "point": (43.3175, 11.3288), "color": "orange"},
            ],
        }
        if png is not None:
            plan["png"] = png
        if base_map is not None:
            plan["base_map"] = base_map
        return plan

    def test_a_google_si_chiede_lo_sfondo_non_la_cartina_finita(self):
        """La richiesta a Google non deve contenere `markers=` né `path=`.

        Non è un vezzo di stile: finché i pallini li disegnava Google, i loro
        colori non coincidevano con quelli della legenda accanto, e con molte
        tappe l'indirizzo superava il limite di lunghezza e la cartina spariva
        del tutto — cioè proprio il difetto fotografato. Se qualcuno domani
        rimette i marker nell'indirizzo, tornano entrambi i problemi."""
        from src import maps_static

        base = maps_static.build_day_base_map_url(self._plan(), "CHIAVE-FINTA")
        self.assertIsNotNone(base, "una giornata con coordinate deve avere il suo sfondo")
        self.assertNotIn("markers=", base["url"])
        self.assertNotIn("path=", base["url"])
        # I parametri di georeferenziazione DEVONO tornare indietro insieme
        # all'indirizzo: senza centro, zoom e scala i nostri pallini non
        # saprebbero in che pixel cadere e finirebbero storti sulle strade.
        for chiave in ("center", "zoom", "size", "scale"):
            self.assertIn(chiave, base)

    def test_lo_sfondo_nudo_non_arriva_mai_al_cliente(self):
        """Se il disegno sopra lo sfondo fallisce, si ripiega sullo schema.

        Consegnare lo sfondo da solo sarebbe il peggiore dei risultati: una
        bella cartina stradale, verosimile, che però non dice NIENTE della
        giornata — e nessuno se ne accorgerebbe guardando in fretta."""
        from src import map_render

        sfondo = self._sfondo()
        # `base_map` incoerente con l'immagine: il disegno sopra non riesce.
        rotto = map_render.attach_local_maps(
            [self._plan(png=b"non-e-un-png", base_map=self._BASE_MAP)]
        )
        self.assertEqual(len(rotto), 1)
        self.assertNotEqual(rotto[0].get("png"), b"non-e-un-png")
        self.assertEqual(rotto[0].get("map_source"), "schema")

        buono = map_render.attach_local_maps(
            [self._plan(png=sfondo, base_map=self._BASE_MAP)]
        )
        self.assertEqual(buono[0].get("map_source"), "google")
        self.assertNotEqual(
            buono[0]["png"], sfondo,
            "lo sfondo è tornato indietro identico: nessuno ci ha disegnato sopra",
        )

    def test_ogni_giornata_esce_con_una_cartina_e_con_la_sua_provenienza(self):
        """Qualunque cosa succeda alla rete, alla chiave o alla quota, ogni
        giornata geolocalizzata esce con una figura, e la figura dichiara da
        dove viene. La provenienza non è un dettaglio interno: decide la frase
        di didascalia sotto la cartina, e una didascalia che promette strade
        dove ci sono solo vettori è una bugia al cliente."""
        from src import map_render

        casi = {
            "sfondo di Google": self._plan(png=self._sfondo(), base_map=self._BASE_MAP),
            "cartina già finita": self._plan(png=self._sfondo()),
            "niente dalla rete": self._plan(),
        }
        for nome, plan in casi.items():
            with self.subTest(caso=nome):
                fuori = map_render.attach_local_maps([plan])[0]
                self.assertTrue(
                    isinstance(fuori.get("png"), bytes) and fuori["png"][:8] == b"\x89PNG\r\n\x1a\n",
                    f"{nome}: la giornata è uscita senza cartina",
                )
                self.assertIn(fuori.get("map_source"), ("google", "schema"))

    def test_la_didascalia_dice_la_verita_su_cosa_si_sta_guardando(self):
        """Le due sorgenti meritano due frasi diverse. Sopra una cartina vera
        si può dire "cartina stradale"; sopra lo schema no, e va detto che le
        strade non ci sono. Il campione locale non ha rete, quindi mostra lo
        schema: qui si controlla che lo dichiari."""
        documento = _leggibile(_documento())
        self.assertIn("Schema in scala delle tappe", documento)
        self.assertNotIn("Cartina stradale della giornata", documento)


class TestLeGuideCoprono(unittest.TestCase):
    """[Richiesta del 2026-08-02: "la parte della guida turistica va
    migliorata: deve esserci una guida per ogni cosa che lo richieda, non
    aver paura di sembrare prolisso è una cosa molto interessante"]"""

    def test_ce_una_guida_per_ogni_tappa_che_la_richiede(self):
        """La regola di Lorenzo non è un numero, è una copertura: "per ogni
        cosa che lo richieda". Chi decide cosa la richiede è
        `guide_generator.select_guide_targets()`, che è la stessa funzione
        usata in produzione — quindi il controllo confronta il documento con
        la selezione vera invece che con una soglia inventata.

        Così, se domani il programma guadagna una tappa e la guida non viene
        scritta, il controllo cade da solo: è il modo per non dover ripetere
        la richiesta."""
        from types import SimpleNamespace

        from src import guide_generator
        from src.pdf_renderer import _slug

        itinerary, _, _, _ = scripts_sample_pdf.build_sample_render_kwargs()
        poi_by_id = {p["id"]: SimpleNamespace(**p) for p in scripts_sample_pdf.POIS}
        attese = guide_generator.select_guide_targets(itinerary, poi_by_id)
        scritte = set(re.findall(r"id='guida-([^']+)' class='anchor-probe'", _documento()))
        # L'ancora nel documento e' `guida-<slug della chiave>`: la chiave
        # grezza ("SAMPLE_campo") non compare mai come tale. Confrontare le
        # chiavi non normalizzate faceva risultare mancanti TUTTE le guide,
        # comprese quelle presenti — un controllo che grida al lupo e che al
        # terzo giro nessuno guarda piu'.
        mancanti = [
            t["name"] for t in attese
            if _slug(t["poi_id"] or t["key"]) not in scritte
        ]
        self.assertEqual(
            [], mancanti,
            f"tappe che meritano una guida e non ce l'hanno: {mancanti}",
        )

    def test_ogni_guida_dice_qualcosa_e_non_solo_il_nome(self):
        """Una scheda con il titolo e sotto due righe generiche è peggio
        della scheda assente: occupa una pagina e non dice niente.

        Il marcatore da cercare è `class='guide-card` SENZA apice di chiusura:
        nel documento la classe è composta (`guide-card page-break`), e
        cercare la forma chiusa non trova nulla — cioè il controllo passerebbe
        a vuoto invece di misurare le schede."""
        html = _documento()
        schede = re.findall(
            r"class='guide-card.*?(?=class='guide-card|class='section-title')",
            html, re.S,
        )
        self.assertTrue(schede, "nessuna scheda di guida trovata nel documento")
        corte = [
            re.sub(r"<[^>]+>", " ", s)[:80].strip()
            for s in schede if len(re.sub(r"<[^>]+>", "", s)) < 300
        ]
        self.assertEqual([], corte, f"schede di guida troppo scarne: {corte}")

    # (marcatore nel documento, cos'è per il cliente)
    BLOCCHI_DI_UNA_GUIDA = [
        ("Cosa cercare, una volta dentro", "cosa guardare davvero sul posto"),
        ("Da sapere", "le curiosità che rendono la scheda una lettura"),
        ("Consigli pratici", "come si visita, non cos'è"),
        ("L'errore che fanno quasi tutti", "l'avvertimento che salva la visita"),
        ("A due passi da qui", "cosa c'è intorno, per non tornare due volte"),
        ("Quando visitare", "l'ora giusta"),
        ("Su misura per te", "il rimando al programma di questo cliente"),
        ("class='disclaimer'", "l'avviso che orari e prezzi vanno verificati"),
    ]

    def test_ogni_scheda_usa_tutti_i_blocchi_che_il_prodotto_sa_scrivere(self):
        """Il campione è il documento su cui si giudica il lavoro: se mostra
        meno di quello che il prodotto emette davvero, fa sembrare povera una
        sezione che povera non è.

        `prompts/system_prompt_guide.txt` chiede a Claude anche `curiosita`,
        `errore_da_evitare`, `dintorni` e `disclaimer`, e il renderer li sa già
        disegnare. Finché il campione non li conteneva, una regressione su uno
        di quei quattro blocchi non l'avrebbe vista nessuno: il documento
        sarebbe uscito identico, perché quei campi non passavano mai di lì.

        Attenzione a una differenza che non si indovina: gli apostrofi dei
        TESTI vengono escapati (`d&#x27;occhio`), quelli delle INTESTAZIONI
        fisse no, perché stanno già scritti nel modello e non passano da
        `_esc()`. Per questo qui "L'errore che fanno quasi tutti" si cerca
        con l'apostrofo vero, mentre altrove serve `_leggibile()`."""
        schede = re.findall(
            r"class='guide-card.*?(?=class='guide-card|class='section-title')",
            _documento(), re.S,
        )
        self.assertTrue(schede, "nessuna scheda di guida trovata nel documento")
        mancanti = sorted({
            f"{marcatore!r} ({motivo})"
            for scheda in schede
            for marcatore, motivo in self.BLOCCHI_DI_UNA_GUIDA
            if marcatore not in scheda
        })
        self.assertEqual(
            [], mancanti,
            f"blocchi che il prodotto sa scrivere e il campione non mostra: {mancanti}",
        )


class TestIlVademecumCopreQuelloCheEStatoChiesto(unittest.TestCase):
    """[Richiesta del 2026-08-02: "aggiungi una parte di vademecum di viaggio
    e di suggerimenti di cosa portare in valigia su come strutturarla, in
    base a dove si va e alla stagione (in base al clima e alle previsioni
    metereologiche) + per eventuali aerei low cost o quando venga richiesto
    quale tipologia di bagaglio conviene prendere (stiva o cabina) e il costo
    di quest'ultimo"] Sono quattro cose distinte, e il controllo le verifica
    una per una: una sezione che ne copre tre su quattro passerebbe un
    controllo generico e fallirebbe la richiesta."""

    ATTESE = ("clima", "valigia", "bagaglio", "cabina", "stiva")

    def test_il_vademecum_parla_di_clima_valigia_e_bagagli(self):
        html = _documento().lower()
        vademecum = html.split("vademecum", 1)[1]
        mancanti = [p for p in self.ATTESE if p not in vademecum]
        self.assertEqual([], mancanti, f"il vademecum non copre: {mancanti}")


class TestLaValigiaSiPuoSpuntare(unittest.TestCase):
    """[Richiesta del 2026-08-02: "per quanto riguarda la parte della valigia
    mi piace molto ma dopo l'elenco vorrei che creassi un collegamento per un
    foglio di calcolo google come quello che ti ho allegato ovviamente
    costruito in base a cio' che richiede la valigia, ma strutturato in
    maniera simile a quello allegato"]

    Il capitolo della valigia e il foglio da spuntare sono due meta' della
    stessa consegna: la lista si LEGGE nel PDF e si SPUNTA nel foglio. Questo
    controllo esiste perche' le due meta' si possono scollare in tre modi, e
    tutti e tre sono silenziosi:

      1. il riquadro sparisce dal PDF (un errore dentro `build_pdf_sections`
         lo fa cadere e il documento esce ugualmente, senza il collegamento);
      2. il riquadro resta ma punta a un file che nella mail non c'e', perche'
         il nome nel riquadro e quello dell'allegato vengono calcolati due
         volte e uno dei due cambia;
      3. il riquadro promette un collegamento che non esiste, perche'
         `CHECKLIST_SHEET_TEMPLATE_URL` e' configurata male e ci si ritrova un
         link vuoto o non cifrato al posto del nome dell'allegato.

    Nessuno dei tre si vede rileggendo il PDF: si vede solo dal cliente che
    clicca e non trova niente."""

    def test_dopo_lelenco_della_valigia_c_e_sempre_il_foglio_da_spuntare(self):
        documento = _leggibile(_documento())
        self.assertIn("Il foglio da spuntare", documento,
                      "il riquadro del foglio di calcolo e' sparito dal capitolo della valigia")

    def test_il_riquadro_sta_nel_capitolo_della_valigia_e_non_altrove(self):
        """Un riquadro giusto nel posto sbagliato e' un riquadro che non si
        legge: va dopo l'elenco, prima di "come si riempie"."""
        documento = _documento()
        self.assertIn("class='vad-sheet'", documento)
        dopo_riquadro = documento.split("class='vad-sheet'", 1)[1]
        self.assertIn("Come si riempie", _leggibile(dopo_riquadro),
                      "il riquadro non e' dentro il vademecum, o e' finito dopo la fine")

    def test_il_nome_scritto_nel_riquadro_e_quello_dell_allegato_vero(self):
        """Il difetto n.2: il riquadro dice "trovi `Valigia-Siena-2026-09.xlsx`
        nella mail" e nella mail c'e' un file che si chiama diversamente."""
        from src import checklist_xlsx
        _, trip, kwargs, _ = scripts_sample_pdf.build_sample_render_kwargs()
        sheet = kwargs.get("checklist_sheet") or {}
        atteso = checklist_xlsx.build_checklist_filename(trip)
        self.assertEqual(atteso, sheet.get("filename"),
                         "il nome nel riquadro non e' quello che genera l'allegato")
        self.assertIn(atteso, _leggibile(_documento()))

    def test_il_foglio_annunciato_esiste_davvero_e_si_apre(self):
        """Il riquadro annuncia N voci: il file allegato deve avere quelle N
        voci, non un numero diverso e non zero."""
        from openpyxl import load_workbook
        from src import checklist_xlsx
        itinerary, trip, kwargs, _ = scripts_sample_pdf.build_sample_render_kwargs()
        annunciate = (kwargs.get("checklist_sheet") or {}).get("rows")
        self.assertTrue(annunciate, "il riquadro annuncia zero voci")
        blob = checklist_xlsx.build_checklist_xlsx(
            trip, kwargs.get("vademecum"), kwargs.get("predeparture"), itinerary)
        self.assertIsInstance(blob, bytes, "il foglio annunciato non si genera")
        import io
        ws = load_workbook(io.BytesIO(blob))["Checklist"]
        # [AGGIORNATO 2026-08-05 — task #193] Prima si contava «tutto dalla
        # riga 2 in giu' tranne le fasce». Da quando sopra l'intestazione ci
        # sono la testata, il pulsante e il contatore, quel conto includeva
        # anche loro. Adesso si conta la cosa che il riquadro del PDF
        # promette davvero: le righe che hanno un'ATTIVITA' scritta, sotto
        # l'intestazione delle colonne.
        intestazione = next(
            r for r in range(1, ws.max_row + 1)
            if ws.cell(row=r, column=1).value == "Priorit\u00e0"
        )
        voci = sum(
            1 for r in range(intestazione + 1, ws.max_row + 1)
            if ws.cell(row=r, column=4).value
        )
        self.assertEqual(annunciate, voci,
                         "il numero di voci promesso nel PDF non e' quello del foglio")

    def test_il_riquadro_non_promette_mai_un_collegamento_che_non_c_e(self):
        """Il difetto n.3. La variabile d'ambiente e' l'unico modo in cui un
        indirizzo puo' entrare qui: se e' vuota o non cifrata, il riquadro
        deve ripiegare sul nome dell'allegato, mai stampare un link a vuoto."""
        from src import pdf_renderer
        casi = {
            "vuota": "",
            "spazi": "   ",
            "non cifrata": "http://esempio.invalido/foglio",
            "non un indirizzo": "chiedimelo-per-mail",
        }
        for nome, url in casi.items():
            with self.subTest(caso=nome):
                html = pdf_renderer._render_checklist_sheet_box({
                    "filename": "Valigia-Siena-2026-09.xlsx", "rows": 28,
                    "url": url if url.startswith("https://") else "",
                    "label": "",
                })
                self.assertIn("Valigia-Siena-2026-09.xlsx", html)
                self.assertNotIn("href=\'\'", html)
                self.assertNotIn("http://", html)
        buono = pdf_renderer._render_checklist_sheet_box({
            "filename": "Valigia-Siena-2026-09.xlsx", "rows": 28,
            "url": "https://docs.google.com/spreadsheets/d/ESEMPIO/edit",
            "label": "Foglio della valigia (Fogli Google)",
        })
        self.assertIn("https://docs.google.com/spreadsheets/d/ESEMPIO/edit", buono)
        self.assertIn("Crea una copia", _leggibile(buono))

    def test_il_riquadro_non_ristampa_lelenco_che_ha_appena_finito(self):
        """La richiesta era un COLLEGAMENTO, non una seconda copia della
        lista: un elenco stampato due volte nello stesso capitolo e' il modo
        piu' rapido di far sembrare gonfio un documento."""
        documento = _documento()
        riquadro = documento.split("class='vad-sheet'", 1)[1].split("</table>", 1)[0]
        self.assertLess(len(riquadro), 1600,
                        "il riquadro si e' allargato: sta ristampando la lista")
        self.assertNotIn("<li", riquadro)


class TestIlRitmoNonLasciaBuchi(unittest.TestCase):
    """[Richiesta del 2026-08-02: "tra le varie attività mi sembra che ci sia
    ancora troppo tempo con il rischio che la gente si annoi oppure finisca
    prima, valuta tu caso per caso ma stacci molto attento"]"""

    def test_il_documento_dichiara_il_ritmo_di_ogni_giornata(self):
        html = _documento()
        self.assertIn("energia", html.lower())

    def test_nessuna_riga_di_spostamento_dice_circa_zero_minuti(self):
        """Difetto reale visto sul PDF: "circa 0 min" stampato sotto una
        tappa. Un dato che non abbiamo non si stampa arrotondato a zero."""
        self.assertNotIn("circa 0 min", _documento())


class TestIlDocumentoNonSiContraddiceNeInventa(unittest.TestCase):
    """Regole che valgono su tutto il documento, sempre, indipendentemente
    dalla richiesta che le ha fatte nascere."""

    def test_nessun_valore_mancante_stampato_come_tale(self):
        html = _documento()
        for veleno in ("None", "nan", "undefined", "[object Object]"):
            self.assertNotIn(
                f">{veleno}<", html,
                f"{veleno!r} finito nel documento del cliente",
            )

    def test_la_copertina_e_la_sintesi_non_dicono_le_stesse_cose(self):
        """[Difetto visto sul PDF vero il 2026-08-02] Due elenchi contigui
        che ripetono destinazione, date, durata e budget: il cliente li legge
        come un errore di stampa, non come una sintesi."""
        html = _documento()
        sintesi = html.split("class='at-a-glance-page'", 1)[1].split("Il viaggio in breve", 1)[0]
        for ripetuto in ("Destinazione", "Durata", "Budget"):
            self.assertNotIn(ripetuto, sintesi)

    def test_nessuna_cella_di_riempimento_vuota_nelle_griglie(self):
        """Una cella vuota della stessa forma delle altre si legge come un
        riquadro che non è stato stampato: un buco sembra un errore, una
        riga chiusa da riquadri più larghi sembra una scelta."""
        html = _documento()
        for griglia in ("cover-facts", "cover-how", "glance-days"):
            if f"class='{griglia}'" not in html:
                continue
            fetta = html.split(f"class='{griglia}'", 1)[1].split("</table>", 1)[0]
            self.assertNotIn("<td></td>", fetta, f"celle vuote in {griglia}")

    def test_i_titoli_dei_capitoli_sono_in_italiano(self):
        """[Difetto visto sul PDF vero il 2026-08-02] Il capitolo di apertura
        si intitolava "Executive Summary" — due parole inglesi in cima alla
        prima pagina di contenuto di un documento che per il resto e' tutto in
        italiano, comprate da un cliente italiano. Non e' un errore di
        sostanza, e' peggio: e' il genere di dettaglio da cui si capisce che
        un documento e' stato assemblato invece che scritto.

        Il campo dei dati continua a chiamarsi `executive_summary`: quello lo
        legge il modello, non il cliente. Qui si guarda solo l'etichetta
        stampata.

        L'unica eccezione ammessa e' il nome proprio di una sezione del
        prodotto — "Architect's Tips" — che infatti si porta dietro la
        traduzione nello stesso titolo. Se domani ne nasce un'altra, va
        aggiunta qui apposta, con la stessa regola: nome inglese ammesso solo
        se il titolo lo spiega in italiano nella stessa riga."""
        titoli = [
            _html.unescape(re.sub("<[^>]+>", "", t)).strip()
            for t in re.findall(r"class='section-title'[^>]*>(.*?)</div>", _documento())
        ]
        self.assertTrue(titoli, "il documento non ha nessun titolo di capitolo")
        # Parole inglesi che in un titolo tradiscono il copia-incolla.
        forestiere = (
            "summary", "executive", "overview", "highlights", "itinerary",
            "budget breakdown", "checklist", "packing", "wrap-up", "insights",
        )
        nomi_propri_ammessi = ("architect's tips",)
        colpevoli = []
        for titolo in titoli:
            minuscolo = titolo.lower()
            if any(nome in minuscolo for nome in nomi_propri_ammessi):
                continue
            for parola in forestiere:
                if parola in minuscolo:
                    colpevoli.append(f"{titolo!r} (contiene {parola!r})")
        self.assertEqual(
            [], colpevoli,
            "titoli di capitolo in inglese in un documento italiano: "
            + "; ".join(colpevoli),
        )

    def test_il_numero_di_giornate_coincide_con_la_durata_dichiarata(self):
        """Un documento che scrive "3 giorni" in copertina e ne elenca due
        nel programma si contraddice sotto gli occhi di chi lo giudica."""
        itinerary, trip, _, _ = scripts_sample_pdf.build_sample_render_kwargs()
        self.assertEqual(trip["duration_days"], len(itinerary["days"]))


class TestIlMotoreDiStampaRegge(unittest.TestCase):
    """Il convertitore HTML→PDF usato in produzione è vecchio e silenzioso:
    davanti a un costrutto che non conosce non protesta, disegna storto. I
    difetti risultanti sono invisibili in HTML e ben visibili sulla carta.
    Questi controlli sono l'unico modo per accorgersene prima del cliente.

    NB: i nomi dei costrutti vietati non vanno mai scritti per esteso in un
    commento di questo file — verrebbero cercati dentro l'HTML e un commento
    li farebbe passare per uso reale. Stanno solo nella lista qui sotto."""

    VIETATI = ("linear-gradient", "display: flex", "display:flex", "rgba(", "opacity:")

    def test_nessun_costrutto_che_il_motore_di_stampa_non_sa_disegnare(self):
        html = _documento()
        trovati = [v for v in self.VIETATI if v in html]
        self.assertEqual(
            [], trovati,
            f"costrutti non supportati dal motore di stampa: {trovati}",
        )

    def test_nessun_riquadro_spuntato_con_caratteri_che_il_font_non_ha(self):
        """I quadratini e le spunte Unicode non esistono nei font del
        motore: sul PDF vero diventano rettangoli vuoti."""
        html = _documento()
        for glifo in ("☐", "☑", "□", "✔"):
            self.assertNotIn(glifo, html)

    def test_nessuna_risorsa_esterna_da_scaricare(self):
        """Nessun CDN, nessun font remoto: il PDF si genera anche quando la
        rete del server non esce, e si genera identico ogni volta."""
        html = _documento()
        self.assertNotIn("<link", html)
        self.assertNotIn("<script", html)


_STRUMENTI = all(shutil.which(x) for x in ("wkhtmltopdf", "pdfinfo", "pdftotext"))


@unittest.skipUnless(_STRUMENTI, "servono wkhtmltopdf e poppler-utils")
class TestLaCartaNonRestaMezzaVuota(unittest.TestCase):
    """[Richiesta del 2026-08-02: "l'impaginazione: troppi spazi vuoti
    dispersivi, migliora"]

    Questo è l'unico difetto che sull'HTML non esiste. Una scheda di guida è
    alta quanto è alta solo dopo che il motore l'ha impaginata: se supera
    metà foglio, due non ci stanno più e il capitolo esce a una scheda per
    pagina con quasi mezza pagina bianca ciascuna. È esattamente quello che
    è successo quando le schede sono passate da due a nove: sette pagine di
    fila piene per poco più della metà. Sull'HTML era invisibile.

    Il controllo non misura i pixel — misura la cosa che al cliente
    interessa: quante pagine costano le schede. Sotto 1,4 schede per pagina
    vuol dire che si è tornati al foglio mezzo vuoto."""

    SCHEDE_PER_PAGINA_MINIME = 1.4
    RIEMPIMENTO_MINIMO = 70.0

    @classmethod
    def setUpClass(cls):
        import scripts_sample_pdf
        from src import pdf_renderer

        itinerary, trip, kwargs, errori = scripts_sample_pdf.build_sample_render_kwargs()
        assert not errori, f"il campione monta con sezioni cadute: {errori}"
        cls._dir = tempfile.TemporaryDirectory()
        cls.pdf = f"{cls._dir.name}/campione.pdf"
        pdf_renderer.render_pdf(itinerary, trip, output_path=cls.pdf, **kwargs)
        cls.attese = len(scripts_sample_pdf.GUIDES)

    @classmethod
    def tearDownClass(cls):
        cls._dir.cleanup()

    def _pagine(self) -> int:
        out = subprocess.run(["pdfinfo", self.pdf], capture_output=True, text=True).stdout
        return int(re.search(r"Pages:\s+(\d+)", out).group(1))

    def _testo_pagina(self, n: int) -> str:
        return subprocess.run(
            ["pdftotext", "-f", str(n), "-l", str(n), self.pdf, "-"],
            capture_output=True, text=True,
        ).stdout

    def test_le_schede_di_guida_non_si_prendono_una_pagina_a_testa(self):
        pagine_con_schede = 0
        schede_viste = 0
        for n in range(1, self._pagine() + 1):
            trovate = self._testo_pagina(n).count("GUIDA TURISTICA TASCABILE")
            if trovate:
                pagine_con_schede += 1
                schede_viste += trovate
        self.assertEqual(
            self.attese, schede_viste,
            "sul PDF non compaiono tutte le schede scritte nel campione",
        )
        densita = schede_viste / pagine_con_schede
        self.assertGreaterEqual(
            densita, self.SCHEDE_PER_PAGINA_MINIME,
            f"le schede occupano {pagine_con_schede} pagine per {schede_viste} guide "
            f"({densita:.2f} per pagina): il capitolo è tornato a stampare "
            f"mezza pagina bianca sotto ogni scheda",
        )

    def test_nessuna_pagina_si_ferma_a_meta_foglio(self):
        """[Richiesta del 2026-08-02: "l'impaginazione: troppi spazi vuoti
        dispersivi, migliora"]

        Il controllo qui sopra guarda un capitolo solo. Questo guarda tutte le
        pagine, perché il bianco si sposta: si stringe la scheda di guida e
        ricompare sotto i piani B, si aggiunge una sezione e ricompare in
        fondo al vademecum. La domanda che si fa il cliente non è "quante
        schede per pagina", è "perché questa pagina è mezza vuota".

        Si misura sull'immagine della pagina, non sul testo: l'ultima riga di
        pixel che porta inchiostro dice fin dove arriva il contenuto, e
        comprende figure, riquadri e cornici, che `pdftotext` non vede.

        L'ULTIMA pagina è esclusa apposta. È la pagina di chiusura: finisce
        dove finisce il documento, e pretendere che arrivi in fondo
        vorrebbe dire riempirla di parole inutili — l'opposto della richiesta.

        La soglia è al 70%: non è "pieno", è "non è mezzo vuoto". Serve
        margine, perché una riga in più o in meno può spostare un blocco di
        pagina, e un controllo che fallisce a ogni virgola smette di essere
        letto."""
        if not shutil.which("pdftoppm"):
            self.skipTest("serve pdftoppm (poppler-utils) per guardare la pagina")
        try:
            import numpy
            from PIL import Image
        except ImportError:
            self.skipTest("servono Pillow e numpy per misurare l'inchiostro")

        with tempfile.TemporaryDirectory() as cartella:
            subprocess.run(
                ["pdftoppm", "-png", "-r", "60", self.pdf, f"{cartella}/pag"],
                check=True, capture_output=True,
            )
            immagini = sorted(pathlib.Path(cartella).glob("pag-*.png"))
            self.assertTrue(immagini, "pdftoppm non ha prodotto nessuna pagina")
            magre = []
            for numero, percorso in enumerate(immagini[:-1], start=1):
                quadro = numpy.array(Image.open(percorso).convert("L"))
                righe_con_inchiostro = numpy.where((quadro < 245).any(axis=1))[0]
                if not len(righe_con_inchiostro):
                    magre.append(f"pagina {numero}: completamente bianca")
                    continue
                arrivo = 100 * righe_con_inchiostro.max() / quadro.shape[0]
                if arrivo < self.RIEMPIMENTO_MINIMO:
                    magre.append(f"pagina {numero}: il contenuto si ferma al {arrivo:.0f}%")
            self.assertEqual(
                [], magre,
                "pagine che restano mezze vuote (l'ultima non conta, è la "
                "chiusura): " + "; ".join(magre),
            )

    def test_ogni_collegamento_interno_diventa_un_salto_vero_sul_PDF(self):
        """[Richiesta del 2026-08-02: "i collegamenti non funzionano: quello
        per la guida turistica che porta in fondo al documento non funziona
        ... non funziona nemmeno il collegamento per le recensioni"]

        `TestICollegamentiPortanoDaQualchePartE` controlla che ogni `href='#x'`
        abbia il suo bersaglio `id='x'` nell'HTML. È necessario ma non basta:
        il cliente non clicca l'HTML, clicca il PDF, e fra i due c'è il motore
        di stampa, che i collegamenti li ricostruisce da capo. Un HTML
        impeccabile può produrre un PDF in cui il salto non c'è — ed è
        successo davvero: il rimando alla guida e quello alla recensione erano
        scritti bene e sulla carta non facevano niente.

        Qui si conta: quanti rimandi interni scrive il documento, e quanti di
        quelli sono diventati un salto che approda su una pagina di QUESTO
        file. I due numeri devono coincidere.

        Un dettaglio che altrimenti si scambia per un errore: nel PDF ogni
        BERSAGLIO lascia a sua volta un'annotazione, larga zero e senza
        destinazione, che serve al motore per sapere dove atterrare. Non è un
        collegamento rotto: è il segnaposto. Si riconosce dal rettangolo di
        area nulla e va contato a parte."""
        try:
            import pypdf
            from pypdf.generic import IndirectObject
        except ImportError:
            self.skipTest("serve pypdf per guardare dentro il PDF")

        attesi = len(re.findall(r"href='#([^']+)'", _documento()))
        self.assertTrue(attesi, "il documento non contiene nessun rimando interno")

        lettore = pypdf.PdfReader(self.pdf)
        pagine_per_id = {
            p.indirect_reference.idnum: i + 1 for i, p in enumerate(lettore.pages)
        }
        saltano, senza_destinazione, fuori = 0, 0, []
        for numero, pagina in enumerate(lettore.pages, start=1):
            annotazioni = pagina.get("/Annots")
            if annotazioni is None:
                continue
            if isinstance(annotazioni, IndirectObject):
                annotazioni = annotazioni.get_object()
            for riferimento in annotazioni:
                annotazione = riferimento.get_object()
                if annotazione.get("/Subtype") != "/Link":
                    continue
                destinazione = annotazione.get("/Dest")
                azione = annotazione.get("/A")
                if azione is not None:
                    azione = azione.get_object()
                    if azione.get("/S") == "/URI":
                        continue  # rimanda fuori dal documento: non è affar nostro
                    if azione.get("/S") == "/GoTo":
                        destinazione = azione.get("/D")
                if destinazione is None:
                    senza_destinazione += 1
                    continue
                if isinstance(destinazione, IndirectObject):
                    destinazione = destinazione.get_object()
                if isinstance(destinazione, (str, bytes)):
                    nominata = lettore.named_destinations.get(str(destinazione))
                    if nominata is None:
                        fuori.append(f"pagina {numero}: nome irrisolto {destinazione!r}")
                        continue
                    bersaglio = nominata.page
                else:
                    bersaglio = destinazione[0]
                if pagine_per_id.get(getattr(bersaglio, "idnum", None)) is None:
                    fuori.append(f"pagina {numero}: destinazione fuori dal documento")
                else:
                    saltano += 1

        self.assertEqual([], fuori, "collegamenti che atterrano nel vuoto: " + "; ".join(fuori))
        self.assertEqual(
            attesi, saltano,
            f"il documento scrive {attesi} rimandi interni ma sulla carta ne "
            f"funzionano {saltano}: qualcuno è tornato a essere solo testo blu",
        )
        bersagli = len(set(re.findall(r"href='#([^']+)'", _documento())))
        self.assertEqual(
            bersagli, senza_destinazione,
            f"i segnaposto di atterraggio sono {senza_destinazione} per "
            f"{bersagli} bersagli distinti: se non coincidono, qualche àncora "
            f"non è stata piazzata dove il rimando la cerca",
        )


# ---------------------------------------------------------------------------
# [AGGIUNTO 2026-08-03 — il giro dello "zoom out dal macro al micro"]
#
# Richiesta di Lorenzo, parola sua: «ora la parola chiave per rinnovare e'
# zoom out dal macro al micro e se hai capito cosa intendo fare hai capito
# l'importanza che hanno le cartine».
#
# Le classi qui sotto difendono quel disegno sul CAMPIONE VERO. Ognuna delle
# richieste di quel giro ha gia' il suo file di controlli dedicato, che pero'
# gira su dati inventati apposta: verificano che la funzione sappia fare la
# cosa. Questi verificano una cosa diversa e piu' fragile — che la cosa sia
# ancora nel documento che parte per la mail. E' esattamente la differenza
# che ha lasciato passare il foglio della valigia senza bottone: il
# costruttore era giusto e provato, nessuno lo chiamava con l'indirizzo.
# ---------------------------------------------------------------------------


class TestLaCartinaSiPuoCliccare(unittest.TestCase):
    """Il macro che porta al micro: e' il perno di tutto il giro.

    L'idea di Lorenzo e' che il documento principale resti magro e che il
    dettaglio (orari, biglietti, come arrivare, guida) si raggiunga
    cliccando il pallino sulla cartina. Se i pallini smettono di essere
    cliccabili non si perde una decorazione: si perde la META' MICRO del
    prodotto, e quello che resta e' un documento piu' povero di prima,
    perche' nel frattempo il dettaglio e' stato tolto dalle pagine.
    """

    def test_le_cartine_hanno_pallini_cliccabili(self):
        self.assertIn(
            "class='map-hit'", _documento(),
            "le cartine non hanno piu' zone cliccabili: il documento e' stato "
            "alleggerito del dettaglio SENZA lasciare il modo di raggiungerlo",
        )

    def test_ogni_pallino_porta_a_un_posto_che_esiste_nel_documento(self):
        """Un pallino che porta nel vuoto e' peggio di un pallino spento.

        Il cliente ci clicca sopra, il lettore PDF non si muove, e lui
        conclude che il documento e' rotto — non che manca una guida.
        """
        documento = _documento()
        ancore = set(re.findall(r"id='([^']+)' class='anchor-probe'", documento))
        rotti = [
            b for b in re.findall(r"<a class='map-hit' href='#([^']+)'", documento)
            if b not in ancore
        ]
        self.assertEqual([], rotti, f"pallini che puntano nel vuoto: {rotti}")

    def test_ogni_pallino_dice_dove_porta_prima_che_ci_si_clicchi(self):
        """Su un PDF il puntatore non cambia forma e non c'e' il "dito".

        Senza il titolo, l'unico modo di sapere che una zona della cartina e'
        cliccabile e' provarci. Il nome del luogo nel `title` e' quello che il
        lettore PDF mostra passandoci sopra: e' l'unico invito che abbiamo.
        """
        senza_nome = [
            m.group(0)[:80]
            for m in re.finditer(r"<a class='map-hit'[^>]*>", _documento())
            if "title='" not in m.group(0) or "title=''" in m.group(0)
        ]
        self.assertEqual([], senza_nome,
                         f"zone cliccabili anonime: {senza_nome}")

    def test_ci_sono_abbastanza_pallini_da_valere_il_meccanismo(self):
        """Due pallini su undici tappe non sono una cartina interattiva.

        La soglia e' bassa apposta — non tutte le tappe hanno una guida — ma
        esiste: senza, il controllo passerebbe con UN solo pallino rimasto in
        piedi per caso, mentre il meccanismo e' di fatto sparito.
        """
        quanti = _documento().count("class='map-hit'")
        self.assertGreaterEqual(
            quanti, 8,
            f"solo {quanti} zone cliccabili sulle cartine dell'intero "
            "documento: il passaggio dal macro al micro non c'e' piu'",
        )

    def test_la_copertina_avverte_che_la_cartina_si_tocca(self):
        """Una funzione che il cliente non sa di avere non e' una funzione.

        Questo controllo e' nato guardando la prima pagina del campione: le
        zone cliccabili c'erano tutte, i collegamenti funzionavano, e da
        nessuna parte era scritto che la cartina si potesse toccare. Su carta
        non c'e' il cursore che cambia forma, quindi il meccanismo restava
        invisibile a chiunque non lo cercasse apposta — cioe' a tutti. Le
        istruzioni "come si legge" in copertina sono l'unico posto dove si
        puo' dire, ed e' li' che va detto.
        """
        # Il marcatore va preso INTERO. Scritto come `find("Come si legge")`
        # questo controllo trovava il COMMENTO nel foglio di stile, 30.000
        # caratteri prima del punto giusto, e falliva su un documento
        # corretto. E' la stessa trappola gia' costata due controlli vacui:
        # il foglio di stile, commenti compresi, fa parte del documento.
        istruzioni = _leggibile(_documento())
        inizio = istruzioni.find("class='cover-how-title'>Come si legge")
        self.assertGreater(inizio, 0, "la copertina non spiega piu' come si legge")
        blocco = istruzioni[inizio:inizio + 1200].lower()
        self.assertIn(
            "cartina si tocca", blocco,
            "la copertina non dice che le tappe sulla cartina sono pulsanti: "
            "il meccanismo c'e' ma il cliente non lo scopre",
        )


class TestIlDettaglioSiRaggiungeMaNonIngombra(unittest.TestCase):
    """«il documento principale appare piu' pulito piu' scarno».

    Alleggerire e' facile: basta togliere. La parte difficile e' togliere
    LASCIANDO il modo di arrivarci, ed e' l'unica meta' che si puo' perdere
    in silenzio — un documento piu' magro sembra un miglioramento anche
    quando e' diventato un documento incompleto.
    """

    def test_come_arrivare_non_e_piu_un_capitolo_a_se(self):
        """«la parte del "come arrivare" appare ridondante, uniscila al
        programma del giorno». Le indicazioni ci sono ancora, ma dentro il
        blocco della tappa: ripetere lo stesso elenco in fondo al documento
        era la ridondanza che Lorenzo ha segnalato."""
        indice = _leggibile(_documento())
        # Si cerca il TITOLO di capitolo, non le parole: "come arrivare"
        # compare legittimamente dentro i blocchi del programma, ed e'
        # proprio dove deve stare adesso.
        self.assertNotIn(
            "class='section-title'>Come arrivare", indice,
            "e' tornato il capitolo separato «Come arrivare»: le stesse "
            "indicazioni due volte nello stesso documento",
        )

    def test_le_indicazioni_di_spostamento_sono_dentro_il_programma(self):
        """L'altra meta' del controllo sopra, e la piu' importante.

        Togliere il capitolo e' un attimo; togliere il capitolo E le
        indicazioni lascia un documento che dice dove andare e non come
        arrivarci. I due controlli vanno letti insieme.
        """
        self.assertIn("class='leg-inline'", _documento(),
                      "le indicazioni di spostamento sono sparite dal "
                      "programma del giorno invece di esserci state spostate")

    def test_ogni_giornata_dice_quanti_chilometri_si_fanno(self):
        """«inserire nel programma del giorno il totale di chilometri/
        percorrenze a piedi». E' l'informazione che decide le scarpe.

        Le giornate si contano sul MARCATORE del titolo, non sulla stringa
        "Giorno N": quella compare anche nell'indice di copertina, nella
        sintesi d'apertura e nei piani B per la pioggia, e contando quelle
        il controllo pretenderebbe piu' totali di quante siano le giornate
        vere — cioe' fallirebbe con il prodotto giusto, che e' il modo piu'
        sicuro di far cancellare un controllo.
        """
        documento = _documento()
        testo = re.sub(r"\s+", " ", _leggibile(re.sub(r"<[^>]+>", " ", documento)))
        giornate = documento.count("<div class='day-title'>")
        totali = len(re.findall(r"In movimento: circa [\d,]+ km", testo))
        self.assertGreater(giornate, 0, "il campione non ha piu' giornate")
        self.assertEqual(
            giornate, totali,
            f"{giornate} giornate ma {totali} totali di percorrenza: qualche "
            "giornata non dice piu' quanto si cammina",
        )

    def test_il_totale_a_piedi_e_scritto_accanto_al_totale_generale(self):
        """Il chilometraggio complessivo comprende bus e treno: quello che
        serve a scegliere le scarpe e' la quota A PIEDI, e senza il minutaggio
        un numero di chilometri non dice niente a chi non cammina di mestiere."""
        testo = _leggibile(re.sub(r"<[^>]+>", " ", _documento()))
        self.assertRegex(
            re.sub(r"\s+", " ", testo),
            r"di cui [\d,]+ km a piedi \(~\d+ min di cammino\)",
            "il totale a piedi (o il tempo di cammino) non compare piu'",
        )


class TestLaGiornataHaUnCriterioDichiarato(unittest.TestCase):
    """«dare un criterio alla programmazione delle cose da vedere
    (minimizzare gli spostamenti, tenendo conto degli orari di apertura
    delle strutture e le varie pause durante la giornata)».

    Lorenzo ha chiesto due cose in una riga, e la seconda si dimentica
    sempre: che il criterio ci sia, e che sia SCRITTO. Un programma
    ottimizzato bene ma senza una riga che lo dica e' indistinguibile da un
    elenco messo in ordine a caso — e il cliente, che il posto non lo
    conosce, non ha nessun modo di accorgersi della differenza.
    """

    CRITERI = [
        ("spostamenti", "Meno spostamenti possibile"),
        ("orari di apertura", "Orari di apertura veri"),
        ("pause", "Le pause sono tappe anche loro"),
    ]

    # Il marcatore va scritto INTERO, virgolette comprese. La trappola e'
    # nota e questo controllo ci e' cascato dentro appena scritto: il foglio
    # di stile fa parte del documento renderizzato, quindi cercare
    # `class='criterio` senza chiudere trova la regola `.criterio { ... }`
    # e dice "verde" anche a riquadro rimosso. Provato: togliendo il
    # riquadro il controllo passava lo stesso.
    MARCATORE = "class='criterio'"

    def test_il_riquadro_del_criterio_c_e(self):
        self.assertIn(self.MARCATORE, _documento(),
                      "il documento non dichiara piu' con che logica e' "
                      "costruita la giornata")

    def test_dichiara_tutti_e_tre_i_criteri_chiesti(self):
        testo = _leggibile(_documento())
        mancanti = [nome for nome, frase in self.CRITERI if frase not in testo]
        self.assertEqual(
            [], mancanti,
            f"criteri chiesti da Lorenzo e non piu' dichiarati: {mancanti}",
        )

    def test_il_criterio_si_dichiara_una_volta_sola(self):
        """Tre righe ripetute a ogni giornata diventano rumore.

        E' il difetto che Lorenzo ha gia' segnalato una volta con altre
        parole («meno testo piu' immagini, non deve essere noioso»): una
        spiegazione utile letta la prima volta e' un ingombro alla quinta.
        """
        quante = _documento().count(self.MARCATORE)
        self.assertLessEqual(
            quante, 1,
            f"il riquadro del criterio compare {quante} volte: ripetuto a "
            "ogni giornata smette di essere una spiegazione e diventa rumore",
        )


class TestOgniGuidaHaLaSuaImmagine(unittest.TestCase):
    """«meno testo piu' immagini, non deve essere noioso».

    In produzione l'immagine e' la FOTOGRAFIA vera del luogo presa da Google
    Places; qui la rete e' chiusa e arriva la copertina disegnata in casa.
    Il controllo verifica quello che vale in tutti e due i casi — che
    un'immagine ci sia, e che la didascalia dica QUALE delle due si sta
    guardando. La didascalia non e' un dettaglio di trasparenza: e' quello
    che impedisce di vendere un disegno spacciandolo per una foto.
    """

    def test_ogni_scheda_di_guida_porta_un_immagine(self):
        documento = _documento()
        schede = documento.count("class='guide-card'")
        immagini = documento.count("class='guide-foto'")
        self.assertGreater(schede, 0, "il campione non ha piu' schede di guida")
        self.assertEqual(
            schede, immagini,
            f"{schede} schede di guida ma {immagini} immagini: il capitolo "
            "e' tornato a essere un muro di testo",
        )

    def test_ogni_immagine_dice_se_e_una_foto_o_un_disegno(self):
        documento = _documento()
        self.assertEqual(
            documento.count("class='guide-foto'"),
            documento.count("class='didascalia'"),
            "qualche immagine e' senza didascalia: il cliente non puo' "
            "distinguere la fotografia del luogo dalla grafica fatta da noi",
        )

    def test_le_immagini_viaggiano_dentro_il_documento(self):
        """Un'immagine richiamata da internet e' un'immagine che un giorno
        non c'e' piu': il PDF resta nella posta del cliente per anni, il
        collegamento no. Tutte le immagini sono incorporate."""
        self.assertNotIn("<img src='http", _documento())


class TestPrimaDiPartireStaInFondo(unittest.TestCase):
    """«la parte del "prima di partire" va messa in fondo al documento».

    Non e' un capriccio di impaginazione. "Prima di partire" e' una lista di
    cose da controllare la sera prima: messa in mezzo interrompe la lettura
    del viaggio con delle incombenze, messa in fondo si ritrova quando
    serve. L'ordine dei capitoli e' pero' la cosa piu' facile da rimettere a
    posto "per sbaglio" aggiungendone uno nuovo.
    """

    @staticmethod
    def _indice_e_corpo():
        """Le due meta' del documento, tagliate sui marcatori veri.

        Tagliare sul primo titolo di capitolo non funziona: i titoli
        compaiono PRIMA nel foglio di stile e nella copertina, e il taglio
        finirebbe a monte dell'indice mettendolo dentro il corpo. Si taglia
        dove finisce l'indice di copertina, che ha un marcatore suo.
        """
        testo = _leggibile(_documento())
        fine = testo.rindex("cover-toc-item")
        return testo[:fine], testo[fine:]

    def test_prima_di_partire_viene_dopo_il_programma_e_dopo_le_guide(self):
        _, corpo = self._indice_e_corpo()
        posizioni = {
            nome: corpo.rindex(nome)
            for nome in ("Il programma, giorno per giorno",
                         "Guide turistiche tascabili", "Prima di partire")
        }
        self.assertLess(posizioni["Il programma, giorno per giorno"],
                        posizioni["Prima di partire"],
                        "«Prima di partire» e' risalito sopra il programma")
        self.assertLess(posizioni["Guide turistiche tascabili"],
                        posizioni["Prima di partire"],
                        "«Prima di partire» e' risalito sopra le guide: "
                        "spezza la lettura del viaggio con delle incombenze")

    def test_lindice_di_copertina_racconta_lo_stesso_ordine_del_corpo(self):
        """Un indice che elenca i capitoli in un ordine diverso da quello in
        cui si trovano e' peggio di nessun indice: si usa proprio quando non
        si vuole sfogliare."""
        indice, corpo = self._indice_e_corpo()
        for prima, dopo in (("Il programma, giorno per giorno", "Prima di partire"),
                            ("Guide turistiche tascabili", "Prima di partire")):
            with self.subTest(coppia=(prima, dopo)):
                self.assertLess(indice.rindex(prima), indice.rindex(dopo),
                                "l'indice elenca i capitoli in un ordine che "
                                "il documento non rispetta")
                self.assertLess(corpo.rindex(prima), corpo.rindex(dopo))


class TestNessunParagrafoSiSpezzaNelCampione(unittest.TestCase):
    """«migliorare l'impaginazione per evitare di spezzare lo stesso
    paragrafo». La regola e' provata a fondo altrove; qui si verifica
    l'unica cosa che quei controlli non possono vedere — che sia applicata
    al documento vero, e non solo alla funzione che sa applicarla."""

    def test_il_campione_passa_dalla_regola_dei_paragrafi(self):
        quanti = _documento().count("<table class='keep-prosa'>")
        self.assertGreater(
            quanti, 10,
            f"solo {quanti} paragrafi protetti dal taglio in tutto il "
            "documento: la passata di impaginazione non gira piu' sul "
            "campione",
        )


if __name__ == "__main__":
    unittest.main()
