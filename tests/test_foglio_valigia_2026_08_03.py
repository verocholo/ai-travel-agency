"""
FOGLIO VALIGIA: il CABLAGGIO del bottone di ritorno — task #184, 2026-08-03.

Richiesta di Lorenzo: «ricordati di aggiungere poi le spunte per i viaggiatori
(se sono tre, 3 caselle di checklist, se sono 4 ne metti 4 e cosi' via) e
ovviamente un pulsante sul foglio di calcolo che ti fa ritornare al pdf
originario».

Perche' un file di controllo NUOVO invece di allungare `test_checklist_xlsx.py`.

Quel file controlla il COSTRUTTORE del foglio, e lo controlla bene: una colonna
per viaggiatore, il tetto delle dodici colonne detto invece che subito, il
bottone su tutti e due i fogli, nessun bottone senza un indirizzo vero. Quando
ho aperto il task #184 quelle funzioni c'erano gia' ed erano gia' verdi.

Il difetto stava tutto ALTROVE: nessuno passava `itinerary_url` al costruttore,
e il campione non passava nemmeno `travellers`. Il foglio sapeva fare le due
cose chieste e non gliele chiedeva nessuno — verde in ogni controllo, e sbagliato
nel file che arriva al cliente. E' la forma di difetto piu' cara del progetto:
la parte che si guarda funziona, quindi non la si guarda piu'.

Questi controlli guardano quindi solo le giunzioni, che sono tre e sono tutte
fragili per motivi diversi:

1. Il MOMENTO. Il foglio si costruisce dentro `build_pdf_extras()`, quando
   l'indirizzo del PDF non esiste ancora — il documento non e' nemmeno stato
   stampato. L'indirizzo nasce dopo, da `publish_hosted_guides()`, che prenota
   il posto prima di scriverci il file. Il foglio va percio' RIFATTO piu' tardi,
   ed e' quello che fa `aggiungi_ritorno_al_foglio_valigia()`.

2. L'ORDINE dentro `service.py`. Il file allegato alla mail si leggeva PRIMA
   della rifacitura: il cliente riceveva lo stesso foglio, e proprio senza la
   cosa che era stata chiesta. E' un difetto invisibile a ogni controllo
   sull'oggetto — il foglio giusto esiste davvero, e' in `sections` — e visibile
   solo guardando in che ordine le due righe stanno scritte.

3. Il CAMPIONE. E' il documento su cui Lorenzo giudica se il prodotto va bene.
   Passava un viaggiatore per difetto e usciva con una colonna di spunte sola:
   mostrava rotta una funzione che rotta non era.
"""
import ast
import io
import pathlib
import unittest

from src import checklist_xlsx
from src import pdf_extras


RADICE = pathlib.Path(__file__).resolve().parent.parent

TRIP = {"destination": "Siena", "date_start": "2026-09-10",
        "date_end": "2026-09-12", "travelers": 3}

VADEMECUM = {
    "climate": {"month_label": "settembre", "zone_label": "mediterraneo"},
    "packing": [{"title": "Sempre", "items": ["scarpe comode", "giacca leggera"]}],
}

PREDEPARTURE = {"checklist": [
    {"title": "Documento d'identita' valido", "detail": "senza non si parte"},
    {"title": "Biglietti del museo stampati", "detail": "la fila si evita cosi'"},
]}

ITINERARIO = {"days": [{
    "day": 1, "title": "Centro", "blocks": [{
        "time": "10:00", "location": "Duomo", "activity": "Visita",
        "duration_min": 60,
    }],
}]}

URL_VERO = "https://ai-travel-agency-service.onrender.com/f/abc/xyz/itinerario.pdf"


def _foglio(travellers=3, itinerary_url=None):
    """Il foglio come lo costruisce il servizio, in memoria."""
    return checklist_xlsx.build_checklist_xlsx(
        TRIP, VADEMECUM, PREDEPARTURE, ITINERARIO,
        travellers=travellers, itinerary_url=itinerary_url,
    )


def _apri(blob):
    import openpyxl

    return openpyxl.load_workbook(io.BytesIO(blob))


def _sezioni(travellers=3):
    """La borsa delle sezioni come esce da `build_pdf_extras()`.

    Il foglio dentro e' quello VERO, costruito senza indirizzo: e' esattamente
    la situazione in cui la rifacitura deve entrare in gioco.
    """
    blob = _foglio(travellers=travellers)
    assert blob, "il foglio di prova e' vuoto: il controllo non misurerebbe niente"
    return {
        "vademecum": VADEMECUM,
        "predeparture": PREDEPARTURE,
        "checklist_xlsx": {
            "filename": "Valigia-Siena-2026-09.xlsx",
            "content": blob,
            "rows": 28,
        },
    }


def _ha_bottone(blob) -> bool:
    wb = _apri(blob)
    return any(
        cella.value == checklist_xlsx.TESTO_BOTTONE_ITINERARIO
        for foglio in wb.worksheets
        for riga in foglio.iter_rows()
        for cella in riga
    )


def _sorgente(nome: str) -> str:
    return (RADICE / nome).read_text(encoding="utf-8")


class TestIlFoglioSiRifaQuandoLIndirizzoFinalmenteEsiste(unittest.TestCase):
    """La giunzione numero 1: il momento."""

    def test_con_un_indirizzo_vero_il_foglio_torna_col_bottone(self):
        """La richiesta di Lorenzo, dal lato di chi la deve cablare."""
        sezioni = _sezioni()
        self.assertFalse(_ha_bottone(sezioni["checklist_xlsx"]["content"]),
                         "il foglio di partenza aveva gia' il bottone: cosi' "
                         "questo controllo passerebbe anche senza la rifacitura")

        fatto = pdf_extras.aggiungi_ritorno_al_foglio_valigia(
            sezioni, URL_VERO, trip=TRIP, itinerary=ITINERARIO, travellers=3,
        )

        self.assertTrue(fatto)
        self.assertTrue(_ha_bottone(sezioni["checklist_xlsx"]["content"]))

    def test_il_bottone_porta_davvero_a_quell_indirizzo(self):
        """Un pulsante che si vede e non apre niente e' peggio di nessuno."""
        sezioni = _sezioni()
        pdf_extras.aggiungi_ritorno_al_foglio_valigia(
            sezioni, URL_VERO, trip=TRIP, itinerary=ITINERARIO, travellers=3,
        )
        wb = _apri(sezioni["checklist_xlsx"]["content"])
        indirizzi = [
            cella.hyperlink.target
            for foglio in wb.worksheets
            for riga in foglio.iter_rows()
            for cella in riga
            if cella.value == checklist_xlsx.TESTO_BOTTONE_ITINERARIO
            and cella.hyperlink is not None
        ]
        self.assertTrue(indirizzi, "il bottone c'e' ma non e' un collegamento")
        for indirizzo in indirizzi:
            self.assertEqual(indirizzo, URL_VERO)

    def test_la_rifacitura_non_perde_il_nome_del_file_ne_il_conteggio(self):
        """Cambia il contenuto, non l'identita' dell'allegato.

        Il nome e il numero di voci finiscono nella mail e nei log. Rifare il
        foglio e restituirlo con un nome diverso vorrebbe dire che il cliente
        riceve un allegato che non e' quello annunciato.
        """
        sezioni = _sezioni()
        prima = dict(sezioni["checklist_xlsx"])
        pdf_extras.aggiungi_ritorno_al_foglio_valigia(
            sezioni, URL_VERO, trip=TRIP, itinerary=ITINERARIO, travellers=3,
        )
        dopo = sezioni["checklist_xlsx"]
        self.assertEqual(dopo["filename"], prima["filename"])
        self.assertEqual(dopo["rows"], prima["rows"])
        self.assertNotEqual(dopo["content"], prima["content"])

    def test_le_colonne_da_spuntare_sopravvivono_alla_rifacitura(self):
        """Il foglio si RIFA' da capo: le spunte vanno richieste di nuovo.

        E' la trappola vera di questa soluzione. Rifare il foglio dimenticando
        `travellers` lo riporterebbe a una colonna sola: si guadagna il bottone
        e si perde, in silenzio, l'altra meta' della richiesta di Lorenzo.
        """
        sezioni = _sezioni(travellers=3)
        pdf_extras.aggiungi_ritorno_al_foglio_valigia(
            sezioni, URL_VERO, trip=TRIP, itinerary=ITINERARIO, travellers=3,
        )
        wb = _apri(sezioni["checklist_xlsx"]["content"])
        ws = wb["Checklist"]
        intestazioni = [
            [c.value for c in riga]
            for riga in ws.iter_rows(max_row=6)
            if "Attivita'" in [c.value for c in riga]
            or "Attività" in [c.value for c in riga]
        ]
        self.assertTrue(intestazioni, "intestazione non trovata nel foglio rifatto")
        spunte = [v for v in intestazioni[0] if isinstance(v, str)
                  and v.startswith("Fatto")]
        self.assertEqual(len(spunte), 3,
                         f"tre viaggiatori, {len(spunte)} colonne da spuntare")


class TestSenzaIndirizzoNonSiToccaNiente(unittest.TestCase):
    """Il foglio senza bottone e' completo e utile: gli manca solo il ritorno.

    Sul portatile e in ogni prova, `PUBLIC_BASE_URL` non c'e' e l'indirizzo e'
    `None`. In quel caso la cosa giusta non e' un bottone finto: e' lasciare
    esattamente il foglio di prima.
    """

    def test_senza_indirizzo_il_foglio_resta_bit_per_bit_quello_di_prima(self):
        sezioni = _sezioni()
        prima = sezioni["checklist_xlsx"]["content"]
        self.assertFalse(pdf_extras.aggiungi_ritorno_al_foglio_valigia(
            sezioni, None, trip=TRIP, itinerary=ITINERARIO, travellers=3,
        ))
        self.assertIs(sezioni["checklist_xlsx"]["content"], prima)

    def test_un_indirizzo_non_sicuro_o_finto_non_diventa_un_bottone(self):
        for indirizzo in ("", "   ", "http://esempio.it/x.pdf", "/tmp/file.pdf",
                          "esempio.it/x.pdf", 42, None):
            with self.subTest(indirizzo=indirizzo):
                sezioni = _sezioni()
                prima = sezioni["checklist_xlsx"]["content"]
                self.assertFalse(pdf_extras.aggiungi_ritorno_al_foglio_valigia(
                    sezioni, indirizzo, trip=TRIP, itinerary=ITINERARIO,
                    travellers=3,
                ))
                self.assertIs(sezioni["checklist_xlsx"]["content"], prima)

    def test_senza_foglio_da_rifare_non_succede_niente_e_non_si_solleva(self):
        """Il foglio e' best-effort: puo' mancare del tutto."""
        for sezioni in ({}, {"checklist_xlsx": None},
                        {"checklist_xlsx": {}},
                        {"checklist_xlsx": {"filename": "x.xlsx"}},
                        {"checklist_xlsx": "non un dizionario"}):
            with self.subTest(sezioni=sezioni):
                self.assertFalse(pdf_extras.aggiungi_ritorno_al_foglio_valigia(
                    sezioni, URL_VERO, trip=TRIP, itinerary=ITINERARIO,
                    travellers=3,
                ))

    def test_ingredienti_rotti_non_possono_togliere_il_pdf_al_cliente(self):
        """Un di piu' non fa mai cadere il prodotto.

        Il documento e' cio' che il cliente ha pagato; il foglio e' un
        accessorio. Se la rifacitura solleva, si porta via l'itinerario intero:
        percio' non deve sollevare per nessun ingrediente.
        """
        sezioni = _sezioni()
        sezioni["vademecum"] = "spazzatura"
        sezioni["predeparture"] = 3.14
        try:
            pdf_extras.aggiungi_ritorno_al_foglio_valigia(
                sezioni, URL_VERO, trip="non un viaggio",
                itinerary=object(), travellers="tre",
            )
        except Exception as e:  # noqa: BLE001 — e' esattamente cio' che si vieta
            self.fail(f"la rifacitura ha sollevato: {type(e).__name__}: {e}")


class TestLAllegatoDellaMailEQuelloCoLBottone(unittest.TestCase):
    """La giunzione numero 2: l'ORDINE delle righe dentro `service.py`.

    Questo e' il difetto vero trovato nel task #184, ed e' l'unico di tutta la
    serie che nessun controllo sull'oggetto poteva vedere: il foglio giusto
    esisteva, era in `sections`, ed era corretto. Semplicemente la riga che lo
    prende per allegarlo alla mail stava PRIMA della riga che lo rifa'. Il
    cliente riceveva la versione senza bottone.

    Si controlla leggendo il sorgente perche' non c'e' altro modo: e' una
    proprieta' dell'ordine in cui le istruzioni sono scritte, non dello stato
    finale del programma.
    """

    def test_il_servizio_legge_il_foglio_dopo_averci_messo_il_bottone(self):
        sorgente = _sorgente("service.py")
        rifacitura = sorgente.find("aggiungi_ritorno_al_foglio_valigia(\n")
        lettura = sorgente.find('checklist_file = sections.get("checklist_xlsx")')
        self.assertGreater(rifacitura, 0, "il servizio non rifa' piu' il foglio")
        self.assertGreater(lettura, 0, "il servizio non allega piu' il foglio")
        self.assertLess(
            rifacitura, lettura,
            "`service.py` prende il foglio da allegare PRIMA di metterci il "
            "bottone di ritorno: alla mail finirebbe la versione senza "
            "bottone, cioe' lo stesso file senza la cosa che era stata chiesta",
        )

    def test_il_foglio_si_legge_prima_del_filtro_a_lista_bianca(self):
        """L'altra trappola dello stesso punto.

        `split_render_kwargs()` tiene solo gli argomenti del renderer, e il
        foglio non e' un argomento del renderer: e' un allegato. Leggerlo dopo
        quel filtro significa non trovarlo piu', e la mail parte senza.
        """
        sorgente = _sorgente("service.py")
        lettura = sorgente.find('checklist_file = sections.get("checklist_xlsx")')
        # Si cerca l'ISTRUZIONE, non il nome: il nome compare prima
        # nell'importazione e dentro i commenti che spiegano proprio questa
        # regola, e cadere su quelli farebbe fallire il controllo mentre il
        # codice e' giusto — il modo migliore per far cancellare un controllo
        # utile perche' "da' sempre fastidio".
        filtro = sorgente.find("sections, section_errors = split_render_kwargs(")
        self.assertGreater(filtro, 0, "il servizio non filtra piu' le sezioni")
        self.assertLess(
            lettura, filtro,
            "il foglio della valigia si legge DOPO il filtro a lista bianca: "
            "li' non c'e' piu', e la mail parte senza allegato",
        )

    def test_anche_la_versione_da_portatile_mette_il_bottone(self):
        """`main.py` e' l'altra strada per lo stesso documento.

        Due strade che producono lo stesso prodotto sono due posti in cui la
        stessa richiesta puo' essere dimenticata. Qui e' successo davvero una
        volta, sul numero di viaggiatori del campione.
        """
        self.assertIn("aggiungi_ritorno_al_foglio_valigia(", _sorgente("main.py"))


class TestIlCampioneMostraLaFunzioneCheEsisteDavvero(unittest.TestCase):
    """La giunzione numero 3: il documento su cui si giudica il prodotto.

    Il campione non e' una prova: e' cio' che Lorenzo guarda per decidere se il
    lavoro e' fatto. Con un viaggiatore per difetto usciva con una colonna sola
    e mostrava rotta una funzione che rotta non era — il modo piu' rapido di
    far rifare due volte lo stesso lavoro.
    """

    def test_il_campione_viaggia_in_piu_di_una_persona(self):
        albero = ast.parse(_sorgente("scripts_sample_pdf.py"))
        viaggiatori = None
        for nodo in ast.walk(albero):
            if isinstance(nodo, ast.Dict):
                for chiave, valore in zip(nodo.keys, nodo.values):
                    if isinstance(chiave, ast.Constant) and \
                            chiave.value == "travelers" and \
                            isinstance(valore, ast.Constant):
                        viaggiatori = valore.value
        self.assertIsNotNone(viaggiatori,
                             "il campione non dice piu' quanti viaggiatori sono")
        self.assertGreater(
            int(viaggiatori), 1,
            "il campione viaggia da solo: le colonne di spunta per viaggiatore "
            "non si vedrebbero nel documento su cui si giudica il prodotto",
        )

    def test_il_campione_passa_davvero_quel_numero_al_foglio(self):
        """Scriverlo nel viaggio e non passarlo e' il difetto che c'era."""
        sorgente = _sorgente("scripts_sample_pdf.py")
        self.assertIn("travellers=int(_trip.get(\"travelers\")", sorgente,
                      "il campione costruisce il foglio senza dire in quanti "
                      "sono: torna a una colonna di spunte sola")

    def test_il_campione_dice_che_il_bottone_di_ritorno_manca_e_perche(self):
        """Onesta' sul campione: il bottone li' non ci puo' essere.

        Porta all'indirizzo pubblico del PDF, che nasce solo quando
        l'itinerario viene ospitato. Sul portatile il PDF e' un file locale.
        Dirlo evita che l'assenza venga scambiata per un difetto — e, peggio,
        che qualcuno la "aggiusti" mettendo un bottone che non apre niente.
        """
        sorgente = _sorgente("scripts_sample_pdf.py")
        self.assertIn("PUBLIC_BASE_URL", sorgente)
        self.assertIn("torna", sorgente.lower())


if __name__ == "__main__":
    unittest.main()
