"""Il design del foglio della valigia (task #193) e l'identità (task #194).

PERCHÉ QUESTO FILE ESISTE
Lorenzo, parola per parola: «ti chiedo di fare la stessa cosa con il foglio di
calcolo (con quest'ultimo svolgi anche un lavoro di miglioramento generale
soprattutto di design)», e più avanti «deve essere facilmente riconoscibile, e
si deve distinguere dal resto del mercato per la sua qualità grafica».

Il design è la cosa che si degrada più facilmente in silenzio: nessun test
tradizionale si accorge se un colore torna al verde predefinito di Fogli
Google, e chi scrive il codice sei mesi dopo non ha visto il documento che
Lorenzo aveva approvato. Questi controlli non giudicano il gusto — non
saprebbero — ma tengono ferme le poche decisioni che rendono il foglio
riconoscibile: che i colori vengano da un posto solo, che ci sia una testata,
che il conto di quello che manca sia VIVO.
"""

import datetime
import io
import unittest

from src import checklist_xlsx, identita


class FakeTrip:
    destination = "Siena"
    date_start = "2026-09-12"
    date_end = "2026-09-15"


VADEMECUM = {
    "climate": {"summary": "Caldo secco, sere fresche.",
                "forecast_link": {"url": "https://esempio.it/meteo",
                                  "label": "Previsioni"}},
    "packing": [{"group": "Abbigliamento",
                 "items": ["Scarpe comode: si cammina su pietra"]},
                {"group": "Documenti e salute",
                 "items": ["Documento d'identità valido per tutta la durata"]}],
    "baggage": {"summary": "Bagaglio a mano sufficiente."},
    "suitcase": {"summary": "Trolley piccolo."},
}
PREDEPARTURE = {"items": [
    {"title": "Controlla la scadenza del documento",
     "detail": "Se scade entro sei mesi, rinnovalo adesso."},
    {"title": "Salva il PDF nei file del telefono"},
]}
ITINERARY = {"days": [
    {"day": 1, "blocks": [{"time": "09:00", "activity": "Salita alla Torre",
                           "location": "Siena"}]},
]}


def _foglio(url=None, viaggiatori=3):
    from openpyxl import load_workbook

    blob = checklist_xlsx.build_checklist_xlsx(
        FakeTrip(), VADEMECUM, PREDEPARTURE, ITINERARY,
        travellers=viaggiatori, itinerary_url=url)
    assert blob, "il foglio non è stato costruito: il resto sarebbe vacuo"
    return load_workbook(io.BytesIO(blob))


def _colonna_a(ws) -> list[str]:
    return [str(c.value or "") for r in ws.iter_rows(min_col=1, max_col=1)
            for c in r]


class TestLIdentitaViveInUnPostoSolo(unittest.TestCase):
    """Tre documenti, una tavolozza.

    I documenti sono diventati tre — l'itinerario, i capitoli staccati e il
    foglio — e un'identità che vive in tre posti smette di essere
    un'identità al primo colore cambiato in due su tre.
    """

    def test_i_colori_del_foglio_vengono_dall_identita(self):
        # Se qualcuno riscrivesse un colore a mano qui dentro, il foglio e il
        # PDF comincerebbero a divergere senza che nessuno se ne accorga.
        self.assertEqual(checklist_xlsx.HEADER_FILL,
                         identita.excel(identita.NOTTE))
        self.assertEqual(checklist_xlsx.BOTTONE_FILL,
                         identita.excel(identita.NOTTE))

    def test_ogni_fascia_ha_il_suo_colore_nell_identita(self):
        for banda in checklist_xlsx.BANDE:
            with self.subTest(banda=banda["key"]):
                self.assertIn(banda["key"], identita.FASCE)
                self.assertEqual(
                    banda["fill"], identita.excel(identita.FASCE[banda["key"]]))

    def test_la_conversione_per_il_foglio_di_calcolo_e_giusta(self):
        # `openpyxl` vuole `AARRGGBB` senza cancelletto: un colore passato
        # nella forma del web viene accettato e poi non salvato — nessun
        # errore, e la cella esce bianca.
        self.assertEqual(identita.excel("#1A3B5C"), "FF1A3B5C")
        self.assertEqual(identita.excel("1a3b5c"), "FF1A3B5C")
        for valore in identita.FASCE.values():
            with self.subTest(valore=valore):
                self.assertRegex(identita.excel(valore), r"^FF[0-9A-F]{6}$")

    def test_ogni_colore_e_una_tinta_piatta_che_il_motore_sa_disegnare(self):
        """L'identità serve anche al PDF, stampato da un motore che non sa
        fare sfumature né trasparenze: un colore con canale alfa sarebbe
        bellissimo nel browser e invisibile nel venduto.

        [SCRITTO DUE VOLTE, 2026-08-05] La prima versione cercava «rgba(»
        dentro il SORGENTE del modulo, e falliva sulla riga di commento che
        spiega perché rgba non si usa. È la trappola che questo progetto ha
        già preso più volte: un controllo sul testo del sorgente vede anche i
        commenti. La versione buona guarda i VALORI, che sono la cosa che
        finisce nel documento."""
        colori = [identita.INCHIOSTRO, identita.NOTTE, identita.ORO,
                  identita.GRIGIO_TESTO, identita.FILETTO, identita.CARTA,
                  identita.AVORIO, *identita.FASCE.values()]
        for colore in colori:
            with self.subTest(colore=colore):
                self.assertRegex(
                    colore, r"^#[0-9A-Fa-f]{6}$",
                    "solo tinte piatte a sei cifre: tutto il resto il motore "
                    "di stampa lo ignora in silenzio",
                )


class TestIlFoglioSiCapisceInDueSecondi(unittest.TestCase):
    """Chi apre il foglio è davanti alla valigia, col telefono in mano."""

    def test_la_prima_riga_dice_di_chi_e_il_documento(self):
        ws = _foglio()["Checklist"]
        self.assertEqual(identita.MARCHIO, ws.cell(row=1, column=1).value)

    def test_la_testata_dice_di_che_viaggio_si_tratta(self):
        testi = " ".join(_colonna_a(_foglio()["Checklist"])[:4])
        self.assertIn("Siena", testi)
        self.assertIn("2026", testi)

    def test_la_testata_e_scura_fino_all_ultima_colonna(self):
        # [REGRESSIONE, la stessa trappola delle fasce] `openpyxl` 3.1.5
        # accetta il colore sulle celle fuse e poi non lo salva: il blocco di
        # testa uscirebbe colorato solo sulla prima colonna. Qui non si
        # fondono celle, e questo controllo è il motivo per cui non si
        # tornerà a fonderle.
        ws = _foglio()["Checklist"]
        primo = ws.cell(row=1, column=1).fill.fgColor.rgb
        ultimo = ws.cell(row=1, column=ws.max_column).fill.fgColor.rgb
        self.assertEqual(primo, identita.excel(identita.NOTTE))
        self.assertEqual(primo, ultimo)

    def test_anche_il_foglio_dell_itinerario_ha_la_sua_testata(self):
        # Due fogli nello stesso file con due facce diverse sono due
        # documenti, non uno.
        ws = _foglio()["Itinerario"]
        self.assertEqual(identita.MARCHIO, ws.cell(row=1, column=1).value)

    def test_le_linguette_sono_colorate(self):
        wb = _foglio()
        self.assertTrue(wb["Checklist"].sheet_properties.tabColor)
        self.assertTrue(wb["Itinerario"].sheet_properties.tabColor)

    def test_la_griglia_di_sfondo_e_spenta(self):
        # Il reticolo grigio è ciò che fa sembrare un foglio di calcolo un
        # foglio di calcolo. Spento, restano i filetti che abbiamo messo noi.
        wb = _foglio()
        self.assertFalse(wb["Checklist"].sheet_view.showGridLines)
        self.assertFalse(wb["Itinerario"].sheet_view.showGridLines)


class TestIlContoDiQuelloCheMancaEVivo(unittest.TestCase):
    """Un numero che non si aggiorna è peggio di nessun numero.

    La prima volta che resta indietro, chi legge smette di fidarsi anche del
    resto del foglio.
    """

    def _riga_contatore(self, ws):
        for riga in _colonna_a(ws):
            if riga.startswith("="):
                return riga
        return ""

    def test_il_contatore_c_e_ed_e_una_formula(self):
        formula = self._riga_contatore(_foglio()["Checklist"])
        self.assertTrue(formula.startswith("="),
                        "il conto è un numero fisso: invecchia al primo tocco")

    def test_conta_le_caselle_non_spuntate(self):
        formula = self._riga_contatore(_foglio()["Checklist"])
        self.assertIn("COUNTIF", formula)
        self.assertIn("FALSE", formula)

    def test_la_formula_e_scritta_in_inglese(self):
        # Dentro il FILE le formule si scrivono sempre in inglese: sono i
        # programmi a mostrarle tradotte. Scrivendo `CONTA.SE` il foglio
        # uscirebbe con un errore `#NOME?` per tutti.
        formula = self._riga_contatore(_foglio()["Checklist"])
        self.assertNotIn("CONTA.SE", formula)

    def test_l_intervallo_copre_tutte_le_colonne_da_spuntare(self):
        # Con tre viaggiatori le colonne sono E, F, G: se l'intervallo si
        # fermasse alla prima, il conto direbbe un terzo del vero.
        ws = _foglio(viaggiatori=3)["Checklist"]
        formula = self._riga_contatore(ws)
        self.assertIn("E", formula)
        self.assertIn("G", formula)

    def test_l_intervallo_parte_dalla_prima_voce_e_non_dall_intestazione(self):
        # Includendo l'intestazione il conto sarebbe sfalsato di una riga per
        # sempre, e nessuno lo noterebbe.
        ws = _foglio()["Checklist"]
        formula = self._riga_contatore(ws)
        intestazione = None
        for indice, valore in enumerate(_colonna_a(ws), start=1):
            if valore == "Priorità":
                intestazione = indice
                break
        self.assertIsNotNone(intestazione)
        self.assertNotIn(f"E{intestazione}:", formula)


class TestLaColonnaPrioritaDiceQualcosa(unittest.TestCase):
    """[REGRESSIONE 2026-08-05] Era una colonna vuota.

    C'era l'intestazione «Priorità» e sotto, per ogni riga, la stringa vuota:
    una colonna di larghezza 11 che non diceva niente e che il filtro non
    poteva usare. Adesso porta la scadenza della fascia, quindi il filtro
    serve: si può chiedere al foglio «fammi vedere solo quello da fare
    subito» attraverso tutte le categorie.
    """

    def test_nessuna_voce_ha_la_priorita_vuota(self):
        ws = _foglio()["Checklist"]
        etichette = {b["label"] for b in checklist_xlsx.BANDE}
        brevi = {b["breve"] for b in checklist_xlsx.BANDE}
        intestazione = None
        vuote = 0
        for riga in ws.iter_rows(min_col=1, max_col=4):
            valore = riga[0].value
            if valore == "Priorità":
                intestazione = riga[0].row
                continue
            if intestazione is None or valore in etichette:
                continue
            if riga[3].value:  # è una voce vera, non una riga di fascia
                if not valore:
                    vuote += 1
                else:
                    self.assertIn(valore, brevi)
        self.assertEqual(vuote, 0, "ci sono voci senza priorità")

    def test_ogni_fascia_ha_la_sua_etichetta_breve(self):
        for banda in checklist_xlsx.BANDE:
            with self.subTest(banda=banda["key"]):
                self.assertTrue(banda.get("breve"))
                # Deve stare nella colonna: undici caratteri, non una frase.
                self.assertLessEqual(len(banda["breve"]), 12)


if __name__ == "__main__":
    unittest.main()
